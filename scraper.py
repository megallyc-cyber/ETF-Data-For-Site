"""
Covered call ETF holdings scraper — Ledger project.

WHERE THIS RUNS
----------------
This script needs outbound access to issuer websites (hamiltonetfs.com,
harvestportfolios.com, bmo.com, globalx.ca, evolveetfs.com, jpmorganfunds.com,
etc). It will NOT run inside a sandboxed dev environment with a restricted
egress allowlist (e.g. one limited to pypi/npm/github). Run it:
  - on your own machine, or
  - as a scheduled GitHub Actions workflow (see .github/workflows/update-data.yml
    example at the bottom of this file), which commits the refreshed JSON back
    to the repo on a cron schedule.

WHAT IT DOES
------------
1. For each fund in FUND_REGISTRY, fetches the issuer's public holdings page
   (most Canadian issuers publish a daily holdings CSV or an HTML table; US
   issuers mostly publish CSV via their fund-data vendor, e.g. SS&C/Confluence).
2. Parses top holdings (ticker, name, weight %).
3. Writes everything to data/funds.json in the schema the website already
   expects (see the FUNDS object in index.html — same shape).

This is a STARTER framework, not a finished, bulletproof scraper: issuer pages
change their markup without notice, some publish only PDFs, and a few only
show sector weights, not per-holding weights. Treat every new issuer as a
small, separate parser function (see the `PARSERS` dict) rather than trying to
write one universal parser — the sites are too different from each other.

Install:
    pip install requests beautifulsoup4 pandas lxml --break-system-packages
"""

import json
import os
from datetime import datetime, timedelta, timezone
import time
import logging
from dataclasses import dataclass, field
from pathlib import Path
from typing import Callable

import re

import requests
from bs4 import BeautifulSoup

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger("ledger-scraper")

OUTPUT_PATH = Path("data/funds.json")
# Some issuers (Hamilton, as of Aug 2026) serve a challenge/interstitial page
# rather than the real one when the request doesn't look like a browser. That
# comes back as a 200 with no holdings table, so it reads as "page structure
# changed" rather than as a block. Sending ordinary browser headers avoids the
# whole category. We still self-identify in the From header and still rate-limit
# ourselves via DELAY_BETWEEN_REQUESTS.
REQUEST_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                  "AppleWebKit/537.36 (KHTML, like Gecko) "
                  "Chrome/126.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,"
              "image/avif,image/webp,*/*;q=0.8",
    "Accept-Language": "en-CA,en;q=0.9",
    "From": "ledger-etf-holdings-aggregator (educational)",
}
# Evolve is the mirror image of Hamilton: it serves holdings happily to a
# self-identified bot but not to browser headers, so headers can't be a single
# global. ACTIVE_HEADERS is swapped per fund by run() before the parser is
# called, which matters because some parsers (Evolve) fetch a second URL —
# their CSV — from inside the parser and must use the same headers.
LEGACY_HEADERS = {
    "User-Agent": "Mozilla/5.0 (compatible; LedgerResearchBot/0.1; "
                  "contact: you@example.com) - educational ETF holdings aggregator"
}
HEADERS_BY_PARSER = {
    "evolve": LEGACY_HEADERS,
}
ACTIVE_HEADERS = REQUEST_HEADERS

REQUEST_TIMEOUT = 20
DELAY_BETWEEN_REQUESTS = 1.5  # be polite, avoid hammering issuer sites


@dataclass
class Fund:
    ticker: str
    name: str
    issuer: str
    region: str          # "CAD" or "US"
    holdings_url: str
    parser: str          # key into PARSERS
    holdings: dict = field(default_factory=dict)  # ticker -> weight_pct
    stats: dict = field(default_factory=dict)     # aum_musd, nav, yield, mer, ...
    distributions: list = field(default_factory=list)  # newest-first payment history
    as_of: str = ""      # UTC date this fund's data was last successfully fetched
    stale: bool = False  # True when we're serving the previous run's data
    fetched_ok: bool = False
    needs_browser: bool = False  # True if the issuer's page requires JS rendering
    # What to wait for once the page is up. Without this a browser fetch
    # just sleeps for a fixed time and hopes, which is how BMO kept
    # returning pages whose holdings table had not been built yet.
    wait_selector: str = None
    # some issuers only render the table after a tab is clicked
    click_selector: str = None
    error: str = ""


# ---------------------------------------------------------------------------
# 1. FUND REGISTRY
# Add every fund you want tracked here. This is the part you'll grow the most
# over time as you find more issuers and more tickers. Keep it data, not code.
# ---------------------------------------------------------------------------
FUND_REGISTRY: list[Fund] = [
    Fund("HDIV", "Hamilton Enhanced Canadian Covered Call ETF", "Hamilton ETFs", "CAD",
         "https://hamiltonetfs.com/etf/hdiv/", "hamilton"),
    Fund("HYLD", "Hamilton Enhanced U.S. Covered Call ETF", "Hamilton ETFs", "CAD",
         "https://hamiltonetfs.com/etf/hyld/", "hamilton"),
    Fund("HMAX", "Hamilton Canadian Financials Yield Maximizer ETF", "Hamilton ETFs", "CAD",
         "https://hamiltonetfs.com/etf/hmax/", "hamilton"),

    Fund("CMAX", "Hamilton Canadian Equity Yield Maximizer ETF", "Hamilton ETFs", "CAD",
         "https://hamiltonetfs.com/etf/cmax/", "hamilton"),
    Fund("IMAX", "Hamilton International Equity Yield Maximizer ETF", "Hamilton ETFs", "CAD",
         "https://hamiltonetfs.com/etf/imax/", "hamilton"),
    Fund("SMAX", "Hamilton U.S. Equity Yield Maximizer ETF", "Hamilton ETFs", "CAD",
         "https://hamiltonetfs.com/etf/smax/", "hamilton"),
    Fund("UMAX", "Hamilton Utilities Yield Maximizer ETF", "Hamilton ETFs", "CAD",
         "https://hamiltonetfs.com/etf/umax/", "hamilton"),
    Fund("QMAX", "Hamilton Technology Yield Maximizer ETF", "Hamilton ETFs", "CAD",
         "https://hamiltonetfs.com/etf/qmax/", "hamilton"),
    Fund("AMAX", "Hamilton Gold Producer Yield Maximizer ETF", "Hamilton ETFs", "CAD",
         "https://hamiltonetfs.com/etf/amax/", "hamilton"),
    Fund("EMAX", "Hamilton Energy Yield Maximizer ETF", "Hamilton ETFs", "CAD",
         "https://hamiltonetfs.com/etf/emax/", "hamilton"),
    Fund("LMAX", "Hamilton Healthcare Yield Maximizer ETF", "Hamilton ETFs", "CAD",
         "https://hamiltonetfs.com/etf/lmax/", "hamilton"),
    Fund("FMAX", "Hamilton U.S. Financials Yield Maximizer ETF", "Hamilton ETFs", "CAD",
         "https://hamiltonetfs.com/etf/fmax/", "hamilton"),
    Fund("RMAX", "Hamilton REITs Yield Maximizer ETF", "Hamilton ETFs", "CAD",
         "https://hamiltonetfs.com/etf/rmax/", "hamilton"),
    Fund("CDAY", "Hamilton Enhanced Canadian Equity DayMAX ETF", "Hamilton ETFs", "CAD",
         "https://hamiltonetfs.com/etf/cday/", "hamilton"),
    Fund("SDAY", "Hamilton Enhanced U.S. Equity DayMAX ETF", "Hamilton ETFs", "CAD",
         "https://hamiltonetfs.com/etf/sday/", "hamilton"),
    Fund("QDAY", "Hamilton Enhanced Technology DayMAX ETF", "Hamilton ETFs", "CAD",
         "https://hamiltonetfs.com/etf/qday/", "hamilton"),
    Fund("BDAY", "Hamilton Enhanced Bitcoin DayMAX ETF", "Hamilton ETFs", "CAD",
         "https://hamiltonetfs.com/etf/bday/", "hamilton"),

    Fund("ZWE", "BMO Europe High Dividend Covered Call Hedged to CAD ETF", "BMO ETFs", "CAD",
         "https://bmogam.com/ca-en/products/exchange-traded-fund/bmo-europe-high-dividend-covered-call-hedged-to-cad-etf-zwe/?tab=holdings", "bmo", needs_browser=True,
         wait_selector="table.holdings", click_selector="span.holdings-sub-section"),
    Fund("ZWG", "BMO Global High Dividend Covered Call ETF", "BMO ETFs", "CAD",
         "https://bmogam.com/ca-en/products/exchange-traded-fund/bmo-global-high-dividend-covered-call-etf-zwg/?tab=holdings", "bmo", needs_browser=True,
         wait_selector="table.holdings", click_selector="span.holdings-sub-section"),
    Fund("ZWH", "BMO US High Dividend Covered Call ETF", "BMO ETFs", "CAD",
         "https://bmogam.com/ca-en/products/exchange-traded-fund/bmo-us-high-dividend-covered-call-etf-zwh/?tab=holdings", "bmo", needs_browser=True,
         wait_selector="table.holdings", click_selector="span.holdings-sub-section"),
    Fund("ZWS", "BMO US High Dividend Covered Call Hedged to CAD ETF", "BMO ETFs", "CAD",
         "https://bmogam.com/ca-en/products/exchange-traded-fund/bmo-us-high-dividend-covered-call-hedged-to-cad-etf-zws/?tab=holdings", "bmo", needs_browser=True,
         wait_selector="table.holdings", click_selector="span.holdings-sub-section"),
    Fund("ZWEN", "BMO Covered Call Energy ETF", "BMO ETFs", "CAD",
         "https://bmogam.com/ca-en/products/exchange-traded-fund/bmo-covered-call-energy-etf-zwen/?tab=holdings", "bmo", needs_browser=True,
         wait_selector="table.holdings", click_selector="span.holdings-sub-section"),
    Fund("ZWHC", "BMO Covered Call Health Care ETF", "BMO ETFs", "CAD",
         "https://bmogam.com/ca-en/products/exchange-traded-fund/bmo-covered-call-health-care-etf-zwhc/?tab=holdings", "bmo", needs_browser=True,
         wait_selector="table.holdings", click_selector="span.holdings-sub-section"),
    Fund("ZWT", "BMO Covered Call Technology ETF", "BMO ETFs", "CAD",
         "https://bmogam.com/ca-en/products/exchange-traded-fund/bmo-covered-call-technology-etf-zwt/?tab=holdings", "bmo", needs_browser=True,
         wait_selector="table.holdings", click_selector="span.holdings-sub-section"),
    Fund("ZWK", "BMO Covered Call US Banks ETF", "BMO ETFs", "CAD",
         "https://bmogam.com/ca-en/products/exchange-traded-fund/bmo-covered-call-us-banks-etf-zwk/?tab=holdings", "bmo", needs_browser=True,
         wait_selector="table.holdings", click_selector="span.holdings-sub-section"),
    Fund("ZPAY", "BMO Premium Yield ETF", "BMO ETFs", "CAD",
         "https://bmogam.com/ca-en/products/exchange-traded-fund/bmo-premium-yield-etf-zpay/?tab=holdings", "bmo", needs_browser=True,
         wait_selector="table.holdings", click_selector="span.holdings-sub-section"),
    Fund("ZWB", "BMO Covered Call Canadian Banks ETF", "BMO ETFs", "CAD",
         "https://bmogam.com/ca-en/products/exchange-traded-fund/bmo-covered-call-canadian-banks-etf-zwb/?tab=holdings", "bmo", needs_browser=True, wait_selector="table.holdings", click_selector="span.holdings-sub-section"),
    Fund("ZWC", "BMO Canadian High Dividend Covered Call ETF", "BMO ETFs", "CAD",
         "https://bmogam.com/ca-en/products/exchange-traded-fund/bmo-canadian-high-dividend-covered-call-etf-zwc/?tab=holdings", "bmo", needs_browser=True, wait_selector="table.holdings", click_selector="span.holdings-sub-section"),
    Fund("ZWU", "BMO Covered Call Utilities ETF", "BMO ETFs", "CAD",
         "https://bmogam.com/ca-en/products/exchange-traded-fund/bmo-covered-call-utilities-etf-zwu/?tab=holdings", "bmo", needs_browser=True, wait_selector="table.holdings", click_selector="span.holdings-sub-section"),
    Fund("ZWP", "BMO Europe High Dividend Covered Call ETF", "BMO ETFs", "CAD",
         "https://bmogam.com/ca-en/products/exchange-traded-fund/bmo-europe-high-dividend-covered-call-etf-zwp/?tab=holdings", "bmo", needs_browser=True, wait_selector="table.holdings", click_selector="span.holdings-sub-section"),

    Fund("HHL", "Harvest Healthcare Leaders Income ETF", "Harvest ETFs", "CAD",
         "https://harvestportfolios.com/etfs/hhl/", "harvest"),
    Fund("HTA", "Harvest Tech Achievers Growth & Income ETF", "Harvest ETFs", "CAD",
         "https://harvestportfolios.com/etf/hta/", "harvest"),
    Fund("HBF", "Harvest US Equity Leaders Income ETF", "Harvest ETFs", "CAD",
         "https://harvestportfolios.com/etf/hbf/", "harvest"),
    Fund("HUTL", "Harvest Utilities Leaders Income ETF", "Harvest ETFs", "CAD",
         "https://harvestportfolios.com/etf/hutl/", "harvest"),
    Fund("HGR", "Harvest REIT Leaders Income ETF", "Harvest ETFs", "CAD",
         "https://harvestportfolios.com/etf/hgr/", "harvest"),
    Fund("HPF", "Harvest Energy Leaders Income ETF", "Harvest ETFs", "CAD",
         "https://harvestportfolios.com/etf/hpf/", "harvest"),
    Fund("HUBL", "Harvest US Bank Leaders Income ETF", "Harvest ETFs", "CAD",
         "https://harvestportfolios.com/etf/hubl/", "harvest"),
    Fund("HLIF", "Harvest Canadian Dividend Leaders Income ETF", "Harvest ETFs", "CAD",
         "https://harvestportfolios.com/etf/hlif/", "harvest"),
    Fund("TRVI", "Harvest Travel & Leisure Income ETF", "Harvest ETFs", "CAD",
         "https://harvestportfolios.com/etf/trvi/", "harvest"),
    Fund("HIND", "Harvest Industrial Leaders Income ETF", "Harvest ETFs", "CAD",
         "https://harvestportfolios.com/etf/hind/", "harvest"),
    Fund("HRIF", "Harvest Diversified Equity Income ETF", "Harvest ETFs", "CAD",
         "https://harvestportfolios.com/etf/hrif/", "harvest"),
    Fund("HPYT", "Harvest Premium Yield Treasury ETF", "Harvest ETFs", "CAD",
         "https://harvestportfolios.com/etf/hpyt/", "harvest"),
    Fund("HPYM", "Harvest Premium Yield 7-10 Year Treasury ETF", "Harvest ETFs", "CAD",
         "https://harvestportfolios.com/etf/hpym/", "harvest"),
    Fund("HBIG", "Harvest Balanced Income Growth ETF", "Harvest ETFs", "CAD",
         "https://harvestportfolios.com/etf/hbig/", "harvest"),
    Fund("HHLE", "Harvest Healthcare Leaders Enhanced Income ETF", "Harvest ETFs", "CAD",
         "https://harvestportfolios.com/etf/hhle/", "harvest"),
    Fund("HTAE", "Harvest Tech Leaders Enhanced Income ETF", "Harvest ETFs", "CAD",
         "https://harvestportfolios.com/etf/htae/", "harvest"),
    Fund("HUTE", "Harvest Utilities Leaders Enhanced Income ETF", "Harvest ETFs", "CAD",
         "https://harvestportfolios.com/etf/hute/", "harvest"),
    Fund("HBIE", "Harvest Balanced Income Growth Enhanced ETF", "Harvest ETFs", "CAD",
         "https://harvestportfolios.com/etf/hbie/", "harvest"),
    Fund("HDIF", "Harvest Diversified Monthly Income ETF", "Harvest ETFs", "CAD",
         "https://harvestportfolios.com/etf/hdif/", "harvest"),
    Fund("HTA", "Harvest Tech Achievers Growth & Income ETF", "Harvest ETFs", "CAD",
         "https://harvestportfolios.com/etf/hta/", "harvest"),

    # Harvest High Income Shares — a separate product suite on its own URL
    # path (/high-income-shares/), which is why the /etf/ registry missed them.
    # Same page structure, so the existing harvest parser handles them as-is.
    Fund("AEME", "Harvest Agnico Eagle Enhanced High Income Shares ETF", "Harvest ETFs", "CAD",
         "https://harvestportfolios.com/high-income-shares/aeme/", "harvest_his"),
    Fund("AMDY", "Harvest AMD Enhanced High Income Shares ETF", "Harvest ETFs", "CAD",
         "https://harvestportfolios.com/high-income-shares/amdy/", "harvest_his"),
    Fund("AMHE", "Harvest Amazon Enhanced High Income Shares ETF", "Harvest ETFs", "CAD",
         "https://harvestportfolios.com/high-income-shares/amhe/", "harvest_his"),
    Fund("AMZH", "Harvest Amazon High Income Shares ETF", "Harvest ETFs", "CAD",
         "https://harvestportfolios.com/high-income-shares/amzh/", "harvest_his"),
    Fund("APLE", "Harvest Apple Enhanced High Income Shares ETF", "Harvest ETFs", "CAD",
         "https://harvestportfolios.com/high-income-shares/aple/", "harvest_his"),
    Fund("AVGY", "Harvest Broadcom Enhanced High Income Shares ETF", "Harvest ETFs", "CAD",
         "https://harvestportfolios.com/high-income-shares/avgy/", "harvest_his"),
    Fund("BCEE", "Harvest BCE Enhanced High Income Shares ETF", "Harvest ETFs", "CAD",
         "https://harvestportfolios.com/high-income-shares/bcee/", "harvest_his"),
    Fund("BLKY", "Harvest Block Enhanced High Income Shares ETF", "Harvest ETFs", "CAD",
         "https://harvestportfolios.com/high-income-shares/blky/", "harvest_his"),
    Fund("CCOE", "Harvest Cameco Enhanced High Income Shares ETF", "Harvest ETFs", "CAD",
         "https://harvestportfolios.com/high-income-shares/ccoe/", "harvest_his"),
    Fund("CNQE", "Harvest CNQ Enhanced High Income Shares ETF", "Harvest ETFs", "CAD",
         "https://harvestportfolios.com/high-income-shares/cnqe/", "harvest_his"),
    Fund("CNYE", "Harvest Coinbase Enhanced High Income Shares ETF", "Harvest ETFs", "CAD",
         "https://harvestportfolios.com/high-income-shares/cnye/", "harvest_his"),
    Fund("CONY", "Harvest Coinbase High Income Shares ETF", "Harvest ETFs", "CAD",
         "https://harvestportfolios.com/high-income-shares/cony/", "harvest_his"),
    Fund("COSY", "Harvest Costco Enhanced High Income Shares ETF", "Harvest ETFs", "CAD",
         "https://harvestportfolios.com/high-income-shares/cosy/", "harvest_his"),
    Fund("CRCY", "Harvest Circle Enhanced High Income Shares ETF", "Harvest ETFs", "CAD",
         "https://harvestportfolios.com/high-income-shares/crcy/", "harvest_his"),
    Fund("CRWY", "Harvest CrowdStrike Enhanced High Income Shares ETF", "Harvest ETFs", "CAD",
         "https://harvestportfolios.com/high-income-shares/crwy/", "harvest_his"),
    Fund("ENBE", "Harvest Enbridge Enhanced High Income Shares ETF", "Harvest ETFs", "CAD",
         "https://harvestportfolios.com/high-income-shares/enbe/", "harvest_his"),
    Fund("GOGY", "Harvest Alphabet Enhanced High Income Shares ETF", "Harvest ETFs", "CAD",
         "https://harvestportfolios.com/high-income-shares/gogy/", "harvest_his"),
    Fund("HHIC", "Harvest Canadian High Income Shares ETF", "Harvest ETFs", "CAD",
         "https://harvestportfolios.com/high-income-shares/hhic/", "harvest_his"),
    Fund("HHIH", "Harvest High Income Equity Shares ETF", "Harvest ETFs", "CAD",
         "https://harvestportfolios.com/high-income-shares/hhih/", "harvest_his"),
    Fund("HHIS", "Harvest Diversified High Income Shares ETF", "Harvest ETFs", "CAD",
         "https://harvestportfolios.com/high-income-shares/hhis/", "harvest_his"),
    Fund("HODY", "Harvest Robinhood Enhanced High Income Shares ETF", "Harvest ETFs", "CAD",
         "https://harvestportfolios.com/high-income-shares/hody/", "harvest_his"),
    Fund("JNJY", "Harvest JnJ Enhanced High Income Shares ETF", "Harvest ETFs", "CAD",
         "https://harvestportfolios.com/high-income-shares/jnjy/", "harvest_his"),
    Fund("JPHE", "Harvest JPHE Enhanced High Income Shares ETF", "Harvest ETFs", "CAD",
         "https://harvestportfolios.com/high-income-shares/jphe/", "harvest_his"),
    Fund("LLHE", "Harvest Eli Lilly Enhanced High Income Shares ETF", "Harvest ETFs", "CAD",
         "https://harvestportfolios.com/high-income-shares/llhe/", "harvest_his"),
    Fund("LLYH", "Harvest Eli Lilly High Income Shares ETF", "Harvest ETFs", "CAD",
         "https://harvestportfolios.com/high-income-shares/llyh/", "harvest_his"),
    Fund("METE", "Harvest Meta Enhanced High Income Shares ETF", "Harvest ETFs", "CAD",
         "https://harvestportfolios.com/high-income-shares/mete/", "harvest_his"),
    Fund("MSFH", "Harvest Microsoft High Income Shares ETF", "Harvest ETFs", "CAD",
         "https://harvestportfolios.com/high-income-shares/msfh/", "harvest_his"),
    Fund("MSHE", "Harvest Microsoft Enhanced High Income Shares ETF", "Harvest ETFs", "CAD",
         "https://harvestportfolios.com/high-income-shares/mshe/", "harvest_his"),
    Fund("MSTE", "Harvest Strategy Inc. Enhanced High Income Shares ETF", "Harvest ETFs", "CAD",
         "https://harvestportfolios.com/high-income-shares/mste/", "harvest_his"),
    Fund("MSTY", "Harvest Strategy Inc. High Income Shares ETF", "Harvest ETFs", "CAD",
         "https://harvestportfolios.com/high-income-shares/msty/", "harvest_his"),
    Fund("NFLY", "Harvest Netflix Enhanced High Income Shares ETF", "Harvest ETFs", "CAD",
         "https://harvestportfolios.com/high-income-shares/nfly/", "harvest_his"),
    Fund("NOVY", "Harvest Novo Enhanced High Income Shares ETF", "Harvest ETFs", "CAD",
         "https://harvestportfolios.com/high-income-shares/novy/", "harvest_his"),
    Fund("NVDH", "Harvest NVIDIA High Income Shares ETF", "Harvest ETFs", "CAD",
         "https://harvestportfolios.com/high-income-shares/nvdh/", "harvest_his"),
    Fund("NVHE", "Harvest NVIDIA Enhanced High Income Shares ETF", "Harvest ETFs", "CAD",
         "https://harvestportfolios.com/high-income-shares/nvhe/", "harvest_his"),
    Fund("ORCY", "Harvest Oracle Enhanced High Income Shares ETF", "Harvest ETFs", "CAD",
         "https://harvestportfolios.com/high-income-shares/orcy/", "harvest_his"),
    Fund("PLTE", "Harvest Palantir Enhanced High Income Shares ETF", "Harvest ETFs", "CAD",
         "https://harvestportfolios.com/high-income-shares/plte/", "harvest_his"),
    Fund("RDDY", "Harvest Reddit Enhanced High Income Shares ETF", "Harvest ETFs", "CAD",
         "https://harvestportfolios.com/high-income-shares/rddy/", "harvest_his"),
    Fund("RYHE", "Harvest Enhanced High Income RY-Linked Shares ETF", "Harvest ETFs", "CAD",
         "https://harvestportfolios.com/high-income-shares/ryhe/", "harvest_his"),
    Fund("SHPE", "Harvest Shopify Enhanced High Income Shares ETF", "Harvest ETFs", "CAD",
         "https://harvestportfolios.com/high-income-shares/shpe/", "harvest_his"),
    Fund("SOFY", "Harvest SoFi Enhanced High Income Shares ETF", "Harvest ETFs", "CAD",
         "https://harvestportfolios.com/high-income-shares/sofy/", "harvest_his"),
    Fund("SPXE", "Harvest SpaceX Enhanced High Income Shares ETF", "Harvest ETFs", "CAD",
         "https://harvestportfolios.com/high-income-shares/spxe/", "harvest_his"),
    Fund("SUHE", "Harvest Suncor Enhanced High Income Shares ETF", "Harvest ETFs", "CAD",
         "https://harvestportfolios.com/high-income-shares/suhe/", "harvest_his"),
    Fund("TDHE", "Harvest Enhanced High Income TD-Linked Shares ETF", "Harvest ETFs", "CAD",
         "https://harvestportfolios.com/high-income-shares/tdhe/", "harvest_his"),
    Fund("TEHE", "Harvest TELUS Enhanced High Income Shares ETF", "Harvest ETFs", "CAD",
         "https://harvestportfolios.com/high-income-shares/tehe/", "harvest_his"),
    Fund("TSLY", "Harvest Tesla Enhanced High Income Shares ETF", "Harvest ETFs", "CAD",
         "https://harvestportfolios.com/high-income-shares/tsly/", "harvest_his"),

    Fund("BANK", "Evolve Canadian Banks and Lifecos Enhanced Yield Fund", "Evolve ETFs", "CAD",
         "https://evolveetfs.com/product/bank/", "evolve", needs_browser=True),
    Fund("CFIN", "Evolve Canadian Financials Yield Fund", "Evolve ETFs", "CAD",
         "https://evolveetfs.com/product/cfin/", "evolve", needs_browser=True),

    Fund("UTES", "Evolve Canadian Utilities Enhanced Yield Index Fund", "Evolve ETFs", "CAD",
         "https://evolveetfs.com/product/utes/", "evolve", needs_browser=True),
    Fund("OILY", "Evolve Canadian Energy Enhanced Yield Index Fund", "Evolve ETFs", "CAD",
         "https://evolveetfs.com/product/oily/", "evolve", needs_browser=True),
    Fund("CUTE", "Evolve Canadian Utilities Yield Fund", "Evolve ETFs", "CAD",
         "https://evolveetfs.com/product/cute/", "evolve", needs_browser=True),
    Fund("ESPX", "Evolve S&P 500 Enhanced Yield Fund", "Evolve ETFs", "CAD",
         "https://evolveetfs.com/product/espx/", "evolve", needs_browser=True),
    Fund("ETSX", "Evolve S&P/TSX 60 Enhanced Yield Fund", "Evolve ETFs", "CAD",
         "https://evolveetfs.com/product/etsx/", "evolve", needs_browser=True),
    Fund("LIFE", "Evolve Global Healthcare Enhanced Yield Fund", "Evolve ETFs", "CAD",
         "https://evolveetfs.com/product/life/", "evolve", needs_browser=True),
    Fund("QQQY", "Evolve NASDAQ Technology Enhanced Yield Index Fund", "Evolve ETFs", "CAD",
         "https://evolveetfs.com/product/qqqy/", "evolve", needs_browser=True),
    Fund("EBNK", "Evolve European Banks Enhanced Yield ETF", "Evolve ETFs", "CAD",
         "https://evolveetfs.com/product/ebnk/", "evolve", needs_browser=True),
    Fund("CALL", "Evolve US Banks Enhanced Yield Fund", "Evolve ETFs", "CAD",
         "https://evolveetfs.com/product/call/", "evolve", needs_browser=True),
    Fund("BASE", "Evolve Global Materials & Mining Enhanced Yield Index ETF", "Evolve ETFs", "CAD",
         "https://evolveetfs.com/product/base/", "evolve", needs_browser=True),
    Fund("LEAD", "Evolve Future Leadership Fund", "Evolve ETFs", "CAD",
         "https://evolveetfs.com/product/lead/", "evolve", needs_browser=True),
    Fund("BOND", "Evolve Enhanced Yield Bond Fund", "Evolve ETFs", "CAD",
         "https://evolveetfs.com/product/bond/", "evolve", needs_browser=True),
    Fund("AGG", "Evolve Canadian Aggregate Bond Enhanced Yield Fund", "Evolve ETFs", "CAD",
         "https://evolveetfs.com/product/agg/", "evolve", needs_browser=True),
    Fund("MIDB", "Evolve Enhanced Yield Mid Term Bond Fund", "Evolve ETFs", "CAD",
         "https://evolveetfs.com/product/midb/", "evolve", needs_browser=True),
    Fund("BIGY", "Evolve US Equity UltraYield ETF", "Evolve ETFs", "CAD",
         "https://evolveetfs.com/product/bigy/", "evolve", needs_browser=True),
    Fund("CANY", "Evolve Canadian Equity UltraYield ETF", "Evolve ETFs", "CAD",
         "https://evolveetfs.com/product/cany/", "evolve", needs_browser=True),
    Fund("SIXY", "Evolve Big Six Canadian Banks UltraYield Index ETF", "Evolve ETFs", "CAD",
         "https://evolveetfs.com/product/sixy/", "evolve", needs_browser=True),
    Fund("INTY", "Evolve International Equity UltraYield ETF", "Evolve ETFs", "CAD",
         "https://evolveetfs.com/product/inty/", "evolve", needs_browser=True),
    Fund("EASY", "Evolve All-in-One UltraYield ETF", "Evolve ETFs", "CAD",
         "https://evolveetfs.com/product/easy/", "evolve", needs_browser=True),
    Fund("TECY", "Evolve NASDAQ Technology UltraYield ETF", "Evolve ETFs", "CAD",
         "https://evolveetfs.com/product/tecy/", "evolve", needs_browser=True),

    Fund("QYLD", "Global X Nasdaq 100 Covered Call ETF", "Global X", "US",
         "https://www.globalxetfs.com/funds/qyld/", "globalx_us"),
    Fund("XYLD", "Global X S&P 500 Covered Call ETF", "Global X", "US",
         "https://www.globalxetfs.com/funds/xyld/", "globalx_us"),
    Fund("RYLD", "Global X Russell 2000 Covered Call ETF", "Global X", "US",
         "https://www.globalxetfs.com/funds/ryld/", "globalx_us"),

    Fund("JEPI", "JPMorgan Equity Premium Income ETF", "J.P. Morgan Asset Management", "US",
         "https://am.jpmorgan.com/us/en/asset-management/adv/products/jpmorgan-equity-premium-income-etf-etf-shares-46641q332", "jpmorgan"),
    Fund("JEPQ", "JPMorgan Nasdaq Equity Premium Income ETF", "J.P. Morgan Asset Management", "US",
         "https://am.jpmorgan.com/us/en/asset-management/adv/products/jpmorgan-nasdaq-equity-premium-income-etf-etf-shares-46654q203", "jpmorgan"),
  

    # Keep adding: Purpose, CI, Brompton, Global X CAD broader lineup,
    # YieldMax (30+ single-stock funds), Simplify, Innovator, NEOS, Amplify...
    Fund("SPXY", "Purpose SpaceX Yield Shares ETF", "Purpose Investments", "CAD",
         "https://www.purposeinvest.com/funds/SPXY-yield-shares-purpose-etf", "purpose_single"),
    Fund("TDY", "Purpose TD Yield Shares ETF", "Purpose Investments", "CAD",
         "https://www.purposeinvest.com/funds/purpose-td-yield-shares-etf", "purpose_single"),
    Fund("RBCY", "Purpose RBC Yield Shares ETF", "Purpose Investments", "CAD",
         "https://www.purposeinvest.com/funds/purpose-rbc-yield-shares-etf", "purpose_single"),
    Fund("BNSY", "Purpose Scotiabank Yield Shares ETF", "Purpose Investments", "CAD",
         "https://www.purposeinvest.com/funds/purpose-scotiabank-yield-shares-etf", "purpose_single"),
    Fund("ENBY", "Purpose Enbridge Yield Shares ETF", "Purpose Investments", "CAD",
         "https://www.purposeinvest.com/funds/purpose-enbridge-yield-shares-etf", "purpose_single"),
    Fund("SHPY", "Purpose Shopify Yield Shares ETF", "Purpose Investments", "CAD",
         "https://www.purposeinvest.com/funds/purpose-shopify-yield-shares-etf", "purpose_single"),
    Fund("CNQY", "Purpose Canadian Natural Resources Yield Shares ETF", "Purpose Investments", "CAD",
         "https://www.purposeinvest.com/funds/purpose-canadian-natural-resources-yield-shares-etf", "purpose_single"),
    Fund("TY", "Purpose TELUS Yield Shares ETF", "Purpose Investments", "CAD",
         "https://www.purposeinvest.com/funds/purpose-telus-yield-shares-etf", "purpose_single"),
    Fund("DOLY", "Purpose Dollarama Yield Shares ETF", "Purpose Investments", "CAD",
         "https://www.purposeinvest.com/funds/purpose-dollarama-yield-shares-etf", "purpose_single"),
    Fund("ATDY", "Purpose Couche-Tard Yield Shares ETF", "Purpose Investments", "CAD",
         "https://www.purposeinvest.com/funds/purpose-couche-tard-yield-shares-etf", "purpose_single"),
    Fund("BNY", "Purpose Brookfield Yield Shares ETF", "Purpose Investments", "CAD",
         "https://www.purposeinvest.com/funds/purpose-brookfield-yield-shares-etf", "purpose_single"),
    Fund("YMAG", "Tech Innovators Yield Shares Purpose ETF", "Purpose Investments", "CAD",
         "https://www.purposeinvest.com/funds/tech-innovators-yield-shares-purpose-etf", "purpose_multi_unsupported"),
    Fund("YCST", "Costco Yield Shares Purpose ETF", "Purpose Investments", "CAD",
         "https://www.purposeinvest.com/funds/costco-yield-shares-purpose-etf", "purpose_single"),
    Fund("YNET", "Netflix Yield Shares Purpose ETF", "Purpose Investments", "CAD",
         "https://www.purposeinvest.com/funds/netflix-yield-shares-purpose-etf", "purpose_single"),
    Fund("YAVG", "Broadcom Yield Shares Purpose ETF", "Purpose Investments", "CAD",
         "https://www.purposeinvest.com/funds/broadcom-yield-shares-purpose-etf", "purpose_single"),
    Fund("YCON", "Coinbase Yield Shares Purpose ETF", "Purpose Investments", "CAD",
         "https://www.purposeinvest.com/funds/coinbase-yield-shares-purpose-etf", "purpose_single"),
    Fund("YPLT", "Palantir Yield Shares Purpose ETF", "Purpose Investments", "CAD",
         "https://www.purposeinvest.com/funds/palantir-yield-shares-purpose-etf", "purpose_single"),
    Fund("YUNH", "UnitedHealth Yield Shares Purpose ETF", "Purpose Investments", "CAD",
         "https://www.purposeinvest.com/funds/unitedhealth-yield-shares-purpose-etf", "purpose_single"),
    Fund("YAMD", "AMD Yield Shares Purpose ETF", "Purpose Investments", "CAD",
         "https://www.purposeinvest.com/funds/amd-yield-shares-purpose-etf", "purpose_single"),
    Fund("YMET", "META Yield Shares Purpose ETF", "Purpose Investments", "CAD",
         "https://www.purposeinvest.com/funds/meta-yield-shares-purpose-etf", "purpose_single"),
    Fund("YNVD", "NVIDIA Yield Shares Purpose ETF", "Purpose Investments", "CAD",
         "https://www.purposeinvest.com/funds/nvidia-yield-shares-purpose-etf", "purpose_single"),
    Fund("MSFY", "Microsoft Yield Shares Purpose ETF", "Purpose Investments", "CAD",
         "https://www.purposeinvest.com/funds/microsoft-yield-shares-purpose-etf", "purpose_single"),
    Fund("YTSL", "Tesla Yield Shares Purpose ETF", "Purpose Investments", "CAD",
         "https://www.purposeinvest.com/funds/tesla-yield-shares-purpose-etf", "purpose_single"),
    Fund("YAMZ", "Amazon Yield Shares Purpose ETF", "Purpose Investments", "CAD",
         "https://www.purposeinvest.com/funds/amazon-yield-shares-purpose-etf", "purpose_single"),
    Fund("APLY", "Apple Yield Shares Purpose ETF", "Purpose Investments", "CAD",
         "https://www.purposeinvest.com/funds/apple-yield-shares-purpose-etf", "purpose_single"),
    Fund("YGOG", "Alphabet Yield Shares Purpose ETF", "Purpose Investments", "CAD",
         "https://www.purposeinvest.com/funds/alphabet-yield-shares-purpose-etf", "purpose_single"),
    Fund("BRKY", "Berkshire Hathaway Yield Shares Purpose ETF", "Purpose Investments", "CAD",
         "https://www.purposeinvest.com/funds/berkshire-hathaway-yield-shares-purpose-etf", "purpose_single"),
    Fund("JEPI", "JPMorgan Equity Premium Income ETF", "J.P. Morgan Asset Management", "US",
         "https://am.jpmorgan.com/FundsMarketingHandler/excel?type=dailyETFHoldings&cusip=46641Q332&country=us&role=adv&fundType=N_ETF&locale=en-US&isUnderlyingHolding=false&isProxyHolding=false", "jpmorgan_xls"),
    Fund("JEPQ", "JPMorgan Nasdaq Equity Premium Income ETF", "J.P. Morgan Asset Management", "US",
         "https://am.jpmorgan.com/FundsMarketingHandler/excel?type=dailyETFHoldings&cusip=46654Q203&country=us&role=adv&fundType=N_ETF&locale=en-US&isUnderlyingHolding=false&isProxyHolding=false", "jpmorgan_xls"),
    Fund("SPYI", "NEOS S&P 500 High Income ETF", "NEOS Investments", "US",
         "https://neosfunds.com/wp-admin/admin-ajax.php?action=download_holdings_csv&ticker=SPYI", "neos_csv"),
    Fund("QQQI", "NEOS Nasdaq-100 High Income ETF", "NEOS Investments", "US",
         "https://neosfunds.com/wp-admin/admin-ajax.php?action=download_holdings_csv&ticker=QQQI", "neos_csv"),
    Fund("IWMI", "NEOS Russell 2000 High Income ETF", "NEOS Investments", "US",
         "https://neosfunds.com/wp-admin/admin-ajax.php?action=download_holdings_csv&ticker=IWMI", "neos_csv"),
    Fund("BTCI", "NEOS Bitcoin High Income ETF", "NEOS Investments", "US",
         "https://neosfunds.com/wp-admin/admin-ajax.php?action=download_holdings_csv&ticker=BTCI", "neos_csv"),
    Fund("QQQH", "NEOS Nasdaq-100 Hedged Equity Income ETF", "NEOS Investments", "US",
         "https://neosfunds.com/wp-admin/admin-ajax.php?action=download_holdings_csv&ticker=QQQH", "neos_csv"),
    Fund("SPYH", "NEOS S&P 500 Hedged Equity Income ETF", "NEOS Investments", "US",
         "https://neosfunds.com/wp-admin/admin-ajax.php?action=download_holdings_csv&ticker=SPYH", "neos_csv"),
    Fund("MLPI", "NEOS MLP & Energy Infrastructure High Income ETF", "NEOS Investments", "US",
         "https://neosfunds.com/wp-admin/admin-ajax.php?action=download_holdings_csv&ticker=MLPI", "neos_csv"),
    Fund("IYRI", "NEOS Real Estate High Income ETF", "NEOS Investments", "US",
         "https://neosfunds.com/wp-admin/admin-ajax.php?action=download_holdings_csv&ticker=IYRI", "neos_csv"),
    Fund("IAUI", "NEOS Gold High Income ETF", "NEOS Investments", "US",
         "https://neosfunds.com/wp-admin/admin-ajax.php?action=download_holdings_csv&ticker=IAUI", "neos_csv"),
    Fund("DIVO", "Amplify CWP Enhanced Dividend Income ETF", "Amplify ETFs", "US",
         "https://amplifyetfs.com/divo-holdings/", "amplify_rendered", needs_browser=True),
    Fund("IDVO", "Amplify CWP International Enhanced Dividend Income ETF", "Amplify ETFs", "US",
         "https://amplifyetfs.com/idvo-holdings/", "amplify_rendered", needs_browser=True),
    Fund("QDVO", "Amplify CWP Growth & Income ETF", "Amplify ETFs", "US",
         "https://amplifyetfs.com/qdvo-holdings/", "amplify_rendered", needs_browser=True),
    Fund("AGCC", "Learn More", "Global X Canada", "CAD",
                  "https://www.globalx.ca/product/agcc#holdings", "globalx_ca_rendered", needs_browser=True),
    Fund("BCCC", "Learn More", "Global X Canada", "CAD",
                  "https://www.globalx.ca/product/bccc#holdings", "globalx_ca_rendered", needs_browser=True),
    Fund("BCCL", "Learn More", "Global X Canada", "CAD",
                  "https://www.globalx.ca/product/bccl#holdings", "globalx_ca_rendered", needs_browser=True),
    Fund("CMCC", "Global X All-In-One Commodity Producers Equity Covered Call ETF", "Global X Canada", "CAD",
                  "https://www.globalx.ca/product/cmcc#holdings", "globalx_ca_rendered", needs_browser=True),
    Fund("CMCL", "Global X Enhanced All-In-One Commodity Producers Equity Covered Call ETF", "Global X Canada", "CAD",
                  "https://www.globalx.ca/product/cmcl#holdings", "globalx_ca_rendered", needs_browser=True),
    Fund("CPCC", "Learn More", "Global X Canada", "CAD",
                  "https://www.globalx.ca/product/cpcc#holdings", "globalx_ca_rendered", needs_browser=True),
    Fund("HGY", "Learn More", "Global X Canada", "CAD",
                  "https://www.globalx.ca/product/hgy#holdings", "globalx_ca_rendered", needs_browser=True),
    Fund("LPAY", "Learn More", "Global X Canada", "CAD",
                  "https://www.globalx.ca/product/lpay#holdings", "globalx_ca_rendered", needs_browser=True),
    Fund("MPAY", "Learn More", "Global X Canada", "CAD",
                  "https://www.globalx.ca/product/mpay#holdings", "globalx_ca_rendered", needs_browser=True),
    Fund("PAYL", "Learn More", "Global X Canada", "CAD",
                  "https://www.globalx.ca/product/payl#holdings", "globalx_ca_rendered", needs_browser=True),
    Fund("PAYM", "Learn More", "Global X Canada", "CAD",
                  "https://www.globalx.ca/product/paym#holdings", "globalx_ca_rendered", needs_browser=True),
    Fund("PAYS", "Learn More", "Global X Canada", "CAD",
                  "https://www.globalx.ca/product/pays#holdings", "globalx_ca_rendered", needs_browser=True),
    Fund("QQCC", "Learn More", "Global X Canada", "CAD",
                  "https://www.globalx.ca/product/qqcc#holdings", "globalx_ca_rendered", needs_browser=True),
    Fund("QQCL", "Learn More", "Global X Canada", "CAD",
                  "https://www.globalx.ca/product/qqcl#holdings", "globalx_ca_rendered", needs_browser=True),
    Fund("RNCC", "Learn More", "Global X Canada", "CAD",
                  "https://www.globalx.ca/product/rncc#holdings", "globalx_ca_rendered", needs_browser=True),
    Fund("RNCL", "Learn More", "Global X Canada", "CAD",
                  "https://www.globalx.ca/product/rncl#holdings", "globalx_ca_rendered", needs_browser=True),
    Fund("RSCC", "Learn More", "Global X Canada", "CAD",
                  "https://www.globalx.ca/product/rscc#holdings", "globalx_ca_rendered", needs_browser=True),
    Fund("RSCL", "Learn More", "Global X Canada", "CAD",
                  "https://www.globalx.ca/product/rscl#holdings", "globalx_ca_rendered", needs_browser=True),
    Fund("SPAY", "Learn More", "Global X Canada", "CAD",
                  "https://www.globalx.ca/product/spay#holdings", "globalx_ca_rendered", needs_browser=True),
    Fund("SVCC", "Learn More", "Global X Canada", "CAD",
                  "https://www.globalx.ca/product/svcc#holdings", "globalx_ca_rendered", needs_browser=True),
    Fund("SVCL", "Global X Enhanced Silver Miners Covered Call ETF", "Global X Canada", "CAD",
                  "https://www.globalx.ca/product/svcl#holdings", "globalx_ca_rendered", needs_browser=True),
    Fund("URCC", "Learn More", "Global X Canada", "CAD",
                  "https://www.globalx.ca/product/urcc#holdings", "globalx_ca_rendered", needs_browser=True),
    Fund("USCC", "Learn More", "Global X Canada", "CAD",
                  "https://www.globalx.ca/product/uscc#holdings", "globalx_ca_rendered", needs_browser=True),
    Fund("USCL", "Learn More", "Global X Canada", "CAD",
                  "https://www.globalx.ca/product/uscl#holdings", "globalx_ca_rendered", needs_browser=True),
    Fund("CNCC", "Global X S&P/TSX 60 Covered Call ETF", "Global X Canada", "CAD",
                  "https://www.globalx.ca/product/cncc#holdings", "globalx_ca_rendered", needs_browser=True),
    Fund("CNCL", "Global X Enhanced S&P/TSX 60 Covered Call ETF", "Global X Canada", "CAD",
                  "https://www.globalx.ca/product/cncl#holdings", "globalx_ca_rendered", needs_browser=True),
    Fund("BKCC", "Global X Equal Weight Canadian Bank Covered Call ETF", "Global X Canada", "CAD",
                  "https://www.globalx.ca/product/bkcc#holdings", "globalx_ca_rendered", needs_browser=True),
    Fund("BKCL", "Global X Enhanced Equal Weight Canadian Banks Covered Call ETF", "Global X Canada", "CAD",
                  "https://www.globalx.ca/product/bkcl#holdings", "globalx_ca_rendered", needs_browser=True),
    Fund("EACC", "Global X MSCI EAFE Covered Call ETF", "Global X Canada", "CAD",
                  "https://www.globalx.ca/product/eacc#holdings", "globalx_ca_rendered", needs_browser=True),
    Fund("EACL", "Global X Enhanced MSCI EAFE Covered Call ETF", "Global X Canada", "CAD",
                  "https://www.globalx.ca/product/eacl#holdings", "globalx_ca_rendered", needs_browser=True),
    Fund("EMCC", "Global X MSCI Emerging Markets Covered Call ETF", "Global X Canada", "CAD",
                  "https://www.globalx.ca/product/emcc#holdings", "globalx_ca_rendered", needs_browser=True),
    Fund("EMCL", "Global X Enhanced MSCI Emerging Markets Covered Call ETF", "Global X Canada", "CAD",
                  "https://www.globalx.ca/product/emcl#holdings", "globalx_ca_rendered", needs_browser=True),
    Fund("ENCC", "Global X Canadian Oil and Gas Equity Covered Call ETF", "Global X Canada", "CAD",
                  "https://www.globalx.ca/product/encc#holdings", "globalx_ca_rendered", needs_browser=True),
    Fund("ENCL", "Global X Enhanced Canadian Oil and Gas Equity Covered Call ETF", "Global X Canada", "CAD",
                  "https://www.globalx.ca/product/encl#holdings", "globalx_ca_rendered", needs_browser=True),
    Fund("EQCC", "Global X All-Equity Asset Allocation Covered Call ETF", "Global X Canada", "CAD",
                  "https://www.globalx.ca/product/eqcc#holdings", "globalx_ca_rendered", needs_browser=True),
    Fund("EQCL", "Global X Enhanced All-Equity Asset Allocation Covered Call ETF", "Global X Canada", "CAD",
                  "https://www.globalx.ca/product/eqcl#holdings", "globalx_ca_rendered", needs_browser=True),
    Fund("GLCC", "Global X Gold Producer Equity Covered Call ETF", "Global X Canada", "CAD",
                  "https://www.globalx.ca/product/glcc#holdings", "globalx_ca_rendered", needs_browser=True),
    Fund("GLCL", "Global X Enhanced Gold Producer Equity Covered Call ETF", "Global X Canada", "CAD",
                  "https://www.globalx.ca/product/glcl#holdings", "globalx_ca_rendered", needs_browser=True),
    Fund("GRCC", "Global X Growth Asset Allocation Covered Call ETF", "Global X Canada", "CAD",
                  "https://www.globalx.ca/product/grcc#holdings", "globalx_ca_rendered", needs_browser=True),
  
    # NEOS funds the registry was missing
    Fund("BNDI", "NEOS Enhanced Income Aggregate Bond ETF", "NEOS Investments", "US",
         "https://neosfunds.com/wp-admin/admin-ajax.php?action=download_holdings_csv&ticker=BNDI", "neos_csv"),
    Fund("CSHI", "NEOS Enhanced Income Cash Alternative ETF", "NEOS Investments", "US",
         "https://neosfunds.com/cshi/", "pdf_holdings"),
    Fund("HYBI", "NEOS Enhanced Income Credit Select ETF", "NEOS Investments", "US",
         "https://neosfunds.com/wp-admin/admin-ajax.php?action=download_holdings_csv&ticker=HYBI", "neos_csv"),
    Fund("NEHI", "NEOS Nasdaq-100 Hedged Equity Income ETF", "NEOS Investments", "US",
         "https://neosfunds.com/wp-admin/admin-ajax.php?action=download_holdings_csv&ticker=NEHI", "neos_csv"),
    Fund("NIHI", "NEOS S&P 500 Hedged Equity Income ETF", "NEOS Investments", "US",
         "https://neosfunds.com/wp-admin/admin-ajax.php?action=download_holdings_csv&ticker=NIHI", "neos_csv"),
    Fund("NLSI", "NEOS Large Cap Systematic Income ETF", "NEOS Investments", "US",
         "https://neosfunds.com/wp-admin/admin-ajax.php?action=download_holdings_csv&ticker=NLSI", "neos_csv"),
    Fund("XBCI", "NEOS Bitcoin High Income ETF", "NEOS Investments", "US",
         "https://neosfunds.com/wp-admin/admin-ajax.php?action=download_holdings_csv&ticker=XBCI", "neos_csv"),
    Fund("XQQI", "NEOS Nasdaq-100 High Income ETF", "NEOS Investments", "US",
         "https://neosfunds.com/wp-admin/admin-ajax.php?action=download_holdings_csv&ticker=XQQI", "neos_csv"),
    Fund("XSPI", "NEOS S&P 500 High Income ETF", "NEOS Investments", "US",
         "https://neosfunds.com/wp-admin/admin-ajax.php?action=download_holdings_csv&ticker=XSPI", "neos_csv"),

    # Broader US option-income universe. Holdings are not parsed for these
    # (each issuer would need its own parser); they carry price history,
    # distributions and a computed yield, and say so on the fund page.
    Fund("GPIX", "Goldman Sachs S&P 500 Core Premium Income ETF", "Goldman Sachs", "US",
         "https://dividendhistory.org/payout/GPIX/", "listing_only"),
    Fund("GPIQ", "Goldman Sachs Nasdaq-100 Core Premium Income ETF", "Goldman Sachs", "US",
         "https://dividendhistory.org/payout/GPIQ/", "listing_only"),
    Fund("XDTE", "Roundhill S&P 500 0DTE Covered Call Strategy ETF", "Roundhill", "US",
         "https://www.roundhillinvestments.com/etf/xdte/", "roundhill", needs_browser=True),
    Fund("QDTE", "Roundhill Innovation-100 0DTE Covered Call Strategy ETF", "Roundhill", "US",
         "https://www.roundhillinvestments.com/etf/qdte/", "roundhill", needs_browser=True),
    Fund("RDTE", "Roundhill Small Cap 0DTE Covered Call Strategy ETF", "Roundhill", "US",
         "https://www.roundhillinvestments.com/etf/rdte/", "roundhill", needs_browser=True),
    Fund("QQQT", "Defiance Nasdaq-100 Enhanced Options Income ETF", "Defiance", "US",
         "https://www.defianceetfs.com/qqqt-full-holdings/", "defiance"),
    Fund("SPYT", "Defiance S&P 500 Enhanced Options Income ETF", "Defiance", "US",
         "https://www.defianceetfs.com/spyt-full-holdings/", "defiance"),
    Fund("CEPI", "REX Crypto Equity Premium Income ETF", "REX Shares", "US",
         "https://www.rexshares.com/cepi/", "rex"),
    Fund("NVII", "REX NVDA Growth & Income ETF", "REX Shares", "US",
         "https://www.rexshares.com/nvii/", "rex"),
    Fund("TSII", "REX TSLA Growth & Income ETF", "REX Shares", "US",
         "https://www.rexshares.com/tsii/", "rex"),
    Fund("WMTI", "REX WMT Growth & Income ETF", "REX Shares", "US",
         "https://www.rexshares.com/wmti/", "rex"),
    Fund("ATCL", "REX Autocallable Income ETF", "REX Shares", "US",
         "https://www.rexshares.com/atcl/", "rex"),
    Fund("DACL", "REX Defensive Autocallable Income ETF", "REX Shares", "US",
         "https://www.rexshares.com/dacl/", "rex"),
    Fund("FEPI", "REX FANG & Innovation Equity Premium Income ETF", "REX Shares", "US",
         "https://www.rexshares.com/fepi/", "rex"),
    Fund("AIPI", "REX AI Equity Premium Income ETF", "REX Shares", "US",
         "https://www.rexshares.com/aipi/", "rex"),
    Fund("BALI", "iShares Advantage Large Cap Income ETF", "iShares", "US",
         "https://dividendhistory.org/payout/BALI/", "listing_only"),
    Fund("TLTW", "iShares 20+ Year Treasury Bond BuyWrite Strategy ETF", "iShares", "US",
         "https://dividendhistory.org/payout/TLTW/", "listing_only"),
    Fund("HYGW", "iShares High Yield Corporate Bond BuyWrite Strategy ETF", "iShares", "US",
         "https://dividendhistory.org/payout/HYGW/", "listing_only"),
    Fund("LQDW", "iShares Investment Grade Corporate Bond BuyWrite Strategy ETF", "iShares", "US",
         "https://dividendhistory.org/payout/LQDW/", "listing_only"),
    Fund("QYLG", "Global X Nasdaq 100 Covered Call & Growth ETF", "Global X", "US",
         "https://www.globalxetfs.com/funds/qylg/", "globalx_us"),
    Fund("XYLG", "Global X S&P 500 Covered Call & Growth ETF", "Global X", "US",
         "https://www.globalxetfs.com/funds/xylg/", "globalx_us"),
    Fund("RYLG", "Global X Russell 2000 Covered Call & Growth ETF", "Global X", "US",
         "https://www.globalxetfs.com/funds/rylg/", "globalx_us"),
    Fund("DJIA", "Global X Dow 30 Covered Call ETF", "Global X", "US",
         "https://www.globalxetfs.com/funds/djia/", "globalx_us"),
    Fund("QRMI", "Global X Nasdaq 100 Risk Managed Income ETF", "Global X", "US",
         "https://www.globalxetfs.com/funds/qrmi/", "globalx_us"),
    Fund("XRMI", "Global X S&P 500 Risk Managed Income ETF", "Global X", "US",
         "https://www.globalxetfs.com/funds/xrmi/", "globalx_us"),
    Fund("ISPY", "ProShares S&P 500 High Income ETF", "ProShares", "US",
         "https://www.proshares.com/our-etfs/strategic/ispy", "proshares"),
    Fund("IQQQ", "ProShares Nasdaq-100 High Income ETF", "ProShares", "US",
         "https://www.proshares.com/our-etfs/strategic/iqqq", "proshares"),
    Fund("JEPY", "Defiance S&P 500 Enhanced Options Income ETF", "Defiance", "US",
         "https://www.defianceetfs.com/jepy-full-holdings/", "defiance"),
    Fund("AIYY", "YieldMax AI Option Income Strategy ETF", "YieldMax", "US",
         "https://www.yieldmaxetfs.com/our-etfs/aiyy/", "yieldmax"),
    Fund("AMDY", "YieldMax AMD Option Income Strategy ETF", "YieldMax", "US",
         "https://www.yieldmaxetfs.com/our-etfs/amdy/", "yieldmax"),
    Fund("APLY", "YieldMax AAPL Option Income Strategy ETF", "YieldMax", "US",
         "https://www.yieldmaxetfs.com/our-etfs/aply/", "yieldmax"),
    Fund("BABO", "YieldMax BABA Option Income Strategy ETF", "YieldMax", "US",
         "https://www.yieldmaxetfs.com/our-etfs/babo/", "yieldmax"),
    Fund("BIGY", "YieldMax Target 12 Big 50 Option Income ETF", "YieldMax", "US",
         "https://www.yieldmaxetfs.com/our-etfs/bigy/", "yieldmax"),
    Fund("BRKC", "YieldMax BRK.B Option Income Strategy ETF", "YieldMax", "US",
         "https://www.yieldmaxetfs.com/our-etfs/brkc/", "yieldmax"),
    Fund("CHPY", "YieldMax Semiconductor Portfolio Option Income ETF", "YieldMax", "US",
         "https://www.yieldmaxetfs.com/our-etfs/chpy/", "yieldmax"),
    Fund("CONY", "YieldMax COIN Option Income Strategy ETF", "YieldMax", "US",
         "https://www.yieldmaxetfs.com/our-etfs/cony/", "yieldmax"),
    Fund("CRCO", "YieldMax CRCL Option Income Strategy ETF", "YieldMax", "US",
         "https://www.yieldmaxetfs.com/our-etfs/crco/", "yieldmax"),
    Fund("CRSH", "YieldMax Short TSLA Option Income Strategy ETF", "YieldMax", "US",
         "https://www.yieldmaxetfs.com/our-etfs/crsh/", "yieldmax"),
    Fund("CVNY", "YieldMax CVNA Option Income Strategy ETF", "YieldMax", "US",
         "https://www.yieldmaxetfs.com/our-etfs/cvny/", "yieldmax"),
    Fund("DDDD", "YieldMax U.S. Stocks Target Double Distribution ETF", "YieldMax", "US",
         "https://www.yieldmaxetfs.com/our-etfs/dddd/", "yieldmax"),
    Fund("DIPS", "YieldMax Short NVDA Option Income Strategy ETF", "YieldMax", "US",
         "https://www.yieldmaxetfs.com/our-etfs/dips/", "yieldmax"),
    Fund("DRAY", "YieldMax DKNG Option Income Strategy ETF", "YieldMax", "US",
         "https://www.yieldmaxetfs.com/our-etfs/dray/", "yieldmax"),
    Fund("FIAT", "YieldMax Short COIN Option Income Strategy ETF", "YieldMax", "US",
         "https://www.yieldmaxetfs.com/our-etfs/fiat/", "yieldmax"),
    Fund("GMEY", "YieldMax GME Option Income Strategy ETF", "YieldMax", "US",
         "https://www.yieldmaxetfs.com/our-etfs/gmey/", "yieldmax"),
    Fund("GPTY", "YieldMax AI & Tech Portfolio Option Income ETF", "YieldMax", "US",
         "https://www.yieldmaxetfs.com/our-etfs/gpty/", "yieldmax"),
    Fund("HIYY", "YieldMax HIMS Option Income Strategy ETF", "YieldMax", "US",
         "https://www.yieldmaxetfs.com/our-etfs/hiyy/", "yieldmax"),
    Fund("HOOY", "YieldMax HOOD Option Income Strategy ETF", "YieldMax", "US",
         "https://www.yieldmaxetfs.com/our-etfs/hooy/", "yieldmax"),
    Fund("INYY", "YieldMax INTC Option Income Strategy ETF", "YieldMax", "US",
         "https://www.yieldmaxetfs.com/our-etfs/inyy/", "yieldmax"),
    Fund("JPO", "YieldMax JP Option Income Strategy ETF", "YieldMax", "US",
         "https://www.yieldmaxetfs.com/our-etfs/jpo/", "yieldmax"),
    Fund("LFGY", "YieldMax Crypto Industry & Tech Portfolio Option Income ETF", "YieldMax", "US",
         "https://www.yieldmaxetfs.com/our-etfs/lfgy/", "yieldmax"),
    Fund("MARO", "YieldMax MARA Option Income Strategy ETF", "YieldMax", "US",
         "https://www.yieldmaxetfs.com/our-etfs/maro/", "yieldmax"),
    Fund("MINY", "YieldMax Strategic Metals & Mining Portfolio Option Income ETF", "YieldMax", "US",
         "https://www.yieldmaxetfs.com/our-etfs/miny/", "yieldmax"),
    Fund("MRNY", "YieldMax MRNA Option Income Strategy ETF", "YieldMax", "US",
         "https://www.yieldmaxetfs.com/our-etfs/mrny/", "yieldmax"),
    Fund("MSST", "YieldMax MSTR Performance & Distribution Target 25 ETF", "YieldMax", "US",
         "https://www.yieldmaxetfs.com/our-etfs/msst/", "yieldmax"),
    Fund("MSTY", "YieldMax MSTR Option Income Strategy ETF", "YieldMax", "US",
         "https://www.yieldmaxetfs.com/our-etfs/msty/", "yieldmax"),
    Fund("NFLY", "YieldMax NFLX Option Income Strategy ETF", "YieldMax", "US",
         "https://www.yieldmaxetfs.com/our-etfs/nfly/", "yieldmax"),
    Fund("NVIT", "YieldMax NVDA Performance & Distribution Target 25 ETF", "YieldMax", "US",
         "https://www.yieldmaxetfs.com/our-etfs/nvit/", "yieldmax"),
    Fund("OARK", "YieldMax Innovation Option Income Strategy ETF", "YieldMax", "US",
         "https://www.yieldmaxetfs.com/our-etfs/oark/", "yieldmax"),
    Fund("PYPY", "YieldMax PYPL Option Income Strategy ETF", "YieldMax", "US",
         "https://www.yieldmaxetfs.com/our-etfs/pypy/", "yieldmax"),
    Fund("QDTY", "YieldMax Nasdaq 100 0DTE Covered Call Strategy ETF", "YieldMax", "US",
         "https://www.yieldmaxetfs.com/our-etfs/qdty/", "yieldmax"),
    Fund("RBLY", "YieldMax RBLX Option Income Strategy ETF", "YieldMax", "US",
         "https://www.yieldmaxetfs.com/our-etfs/rbly/", "yieldmax"),
    Fund("RDTY", "YieldMax R2000 0DTE Covered Call Strategy ETF", "YieldMax", "US",
         "https://www.yieldmaxetfs.com/our-etfs/rdty/", "yieldmax"),
    Fund("RDYY", "YieldMax RDDT Option Income Strategy ETF", "YieldMax", "US",
         "https://www.yieldmaxetfs.com/our-etfs/rdyy/", "yieldmax"),
    Fund("RNTY", "YieldMax Target 12 Real Estate Option Income ETF", "YieldMax", "US",
         "https://www.yieldmaxetfs.com/our-etfs/rnty/", "yieldmax"),
    Fund("SDTY", "YieldMax S&P 500 0DTE Covered Call Strategy ETF", "YieldMax", "US",
         "https://www.yieldmaxetfs.com/our-etfs/sdty/", "yieldmax"),
    Fund("SLTY", "YieldMax Ultra Short Option Income Strategy ETF", "YieldMax", "US",
         "https://www.yieldmaxetfs.com/our-etfs/slty/", "yieldmax"),
    Fund("SNOY", "YieldMax SNOW Option Income Strategy ETF", "YieldMax", "US",
         "https://www.yieldmaxetfs.com/our-etfs/snoy/", "yieldmax"),
    Fund("SOXY", "YieldMax Target 12 Semiconductor Option Income ETF", "YieldMax", "US",
         "https://www.yieldmaxetfs.com/our-etfs/soxy/", "yieldmax"),
    Fund("TSLY", "YieldMax TSLA Option Income Strategy ETF", "YieldMax", "US",
         "https://www.yieldmaxetfs.com/our-etfs/tsly/", "yieldmax"),
    Fund("TSMY", "YieldMax TSM Option Income Strategy ETF", "YieldMax", "US",
         "https://www.yieldmaxetfs.com/our-etfs/tsmy/", "yieldmax"),
    Fund("WNTR", "YieldMax MSTR Short Option Income Strategy ETF", "YieldMax", "US",
         "https://www.yieldmaxetfs.com/our-etfs/wntr/", "yieldmax"),
    Fund("XOMO", "YieldMax XOM Option Income Strategy ETF", "YieldMax", "US",
         "https://www.yieldmaxetfs.com/our-etfs/xomo/", "yieldmax"),
    Fund("XYZY", "YieldMax XYZ Option Income Strategy ETF", "YieldMax", "US",
         "https://www.yieldmaxetfs.com/our-etfs/xyzy/", "yieldmax"),
    Fund("YBIT", "YieldMax Bitcoin Option Income Strategy ETF", "YieldMax", "US",
         "https://www.yieldmaxetfs.com/our-etfs/ybit/", "yieldmax"),
    Fund("YMAG", "YieldMax Magnificent 7 Fund of Option Income ETF", "YieldMax", "US",
         "https://www.yieldmaxetfs.com/our-etfs/ymag/", "yieldmax"),
    Fund("YQQQ", "YieldMax Short N100 Option Income Strategy ETF", "YieldMax", "US",
         "https://www.yieldmaxetfs.com/our-etfs/yqqq/", "yieldmax"),
    Fund("YRAM", "YieldMax Memory and Storage Portfolio Option Income ETF", "YieldMax", "US",
         "https://www.yieldmaxetfs.com/our-etfs/yram/", "yieldmax"),
    Fund("YSPC", "YieldMax SPCX Option Income Strategy ETF", "YieldMax", "US",
         "https://www.yieldmaxetfs.com/our-etfs/yspc/", "yieldmax"),
    Fund("ULTY", "YieldMax Ultra Option Income Strategy ETF", "YieldMax", "US",
         "https://www.yieldmaxetfs.com/our-etfs/ulty/", "yieldmax"),
    Fund("YMAX", "YieldMax Universe Fund of Option Income ETFs", "YieldMax", "US",
         "https://www.yieldmaxetfs.com/our-etfs/ymax/", "yieldmax"),
    Fund("NVDY", "YieldMax NVDA Option Income Strategy ETF", "YieldMax", "US",
         "https://www.yieldmaxetfs.com/our-etfs/nvdy/", "yieldmax"),
    Fund("AMZY", "YieldMax AMZN Option Income Strategy ETF", "YieldMax", "US",
         "https://www.yieldmaxetfs.com/our-etfs/amzy/", "yieldmax"),
    Fund("GOOY", "YieldMax GOOGL Option Income Strategy ETF", "YieldMax", "US",
         "https://www.yieldmaxetfs.com/our-etfs/gooy/", "yieldmax"),
    Fund("FBY", "YieldMax META Option Income Strategy ETF", "YieldMax", "US",
         "https://www.yieldmaxetfs.com/our-etfs/fby/", "yieldmax"),
    Fund("MSFO", "YieldMax MSFT Option Income Strategy ETF", "YieldMax", "US",
         "https://www.yieldmaxetfs.com/our-etfs/msfo/", "yieldmax"),
    Fund("PLTY", "YieldMax PLTR Option Income Strategy ETF", "YieldMax", "US",
         "https://www.yieldmaxetfs.com/our-etfs/plty/", "yieldmax"),
    Fund("SMCY", "YieldMax SMCI Option Income Strategy ETF", "YieldMax", "US",
         "https://www.yieldmaxetfs.com/our-etfs/smcy/", "yieldmax"),
    Fund("GDXY", "YieldMax Gold Miners Option Income Strategy ETF", "YieldMax", "US",
         "https://www.yieldmaxetfs.com/our-etfs/gdxy/", "yieldmax"),

]


# ---------------------------------------------------------------------------

def parse_purpose_single(html: str) -> dict:
    """Purpose Yield Shares are single-stock covered call funds — holdings are
    essentially 100% the underlying stock plus written call options. The
    underlying ticker appears in the page title in parentheses, e.g.
    'Apple (AAPL) Yield Shares Purpose ETF'."""
    import re
    head_end = html.find("<body")
    head_html = html[:head_end] if head_end != -1 else html
    match = re.search(r"\(([A-Z]{1,5})\)\s+Yield Shares", head_html)
    if not match:
        raise ValueError("could not find underlying ticker in Purpose fund page title")
    return {match.group(1): 100.0}


# 2. PER-ISSUER PARSERS
# Each issuer publishes holdings differently. Write one small function per
# issuer. Return {ticker: weight_pct}. Keep these easy to fix in isolation —
# when an issuer changes their page, only their parser breaks, not everything.
# ---------------------------------------------------------------------------

def parse_hamilton(html: str) -> dict:
    """Hamilton ETFs typically render a holdings table with columns:
    Ticker | Security | Weight (%). Adjust selectors after inspecting the
    live page — this is a starting guess at the structure."""
    soup = BeautifulSoup(html, "lxml")
    holdings = {}
    table = soup.select_one('table[id^="etf-holdings-"]')
    if not table:
        raise ValueError("holdings table not found — page structure may have changed")
    rows = table.select("tr")
    for row in rows[1:]:
        cells = [c.get_text(strip=True) for c in row.select("td")]
        if len(cells) < 3:
            continue
        ticker, weight = cells[0], cells[2]
        if not ticker or not weight:
            continue
        try:
            w = float(weight.replace("%", "").replace(",", ""))
        except ValueError:
            continue
        if w <= 0:
            continue
        holdings[ticker] = w
    return holdings

HIS_NAME_TO_TICKER = {
    "nvidia": "NVDA", "tesla": "TSLA", "apple": "AAPL", "amazon": "AMZN",
    "microsoft": "MSFT", "alphabet": "GOOGL", "google": "GOOGL", "meta": "META",
    "advanced micro": "AMD", "palantir": "PLTR", "coinbase": "COIN",
    "microstrategy": "MSTR", "strategy inc": "MSTR", "netflix": "NFLX",
    "broadcom": "AVGO", "eli lilly": "LLY", "johnson & johnson": "JNJ",
    "jpmorgan": "JPM", "costco": "COST", "berkshire": "BRK.B", "shopify": "SHOP",
    "royal bank": "RY", "toronto-dominion": "TD", "bce": "BCE", "enbridge": "ENB",
    "canadian natural": "CNQ", "suncor": "SU", "telus": "T", "bank of montreal": "BMO",
    "bank of nova scotia": "BNS", "agnico": "AEM", "constellation": "CSU",
    "reddit": "RDDT", "novo nordisk": "NVO", "oracle": "ORCL", "crowdstrike": "CRWD",
    "circle": "CRCL", "sofi": "SOFI", "blackrock": "BLK", "spacex": "SPACEX",
    "cameco": "CCO", "hood": "HOOD", "robinhood": "HOOD", "cenovus": "CVE",
}


def parse_harvest_his(html: str) -> dict:
    """Harvest High Income Shares list a single underlying by company NAME.

    Their /high-income-shares/ pages differ from the /etf/ pages: the holdings
    table is headed HOLDING and rows read like 'NVIDIA Corporation | 126.0%'.
    Weights exceed 100% because these funds use leverage — that is real, not a
    parsing artefact, so it is kept as published. Names are resolved to tickers
    where we recognise them so the overlap tool can match them against other
    funds; unrecognised names are kept verbatim rather than dropped.
    """
    soup = BeautifulSoup(html, "lxml")
    holdings = {}
    for table in soup.find_all("table"):
        head = table.find("tr")
        if not head:
            continue
        first = head.find(["th", "td"])
        if not first or "holding" not in first.get_text(strip=True).lower():
            continue
        for tr in table.find_all("tr")[1:]:
            cells = [c.get_text(" ", strip=True) for c in tr.find_all(["td", "th"])]
            if len(cells) < 2:
                continue
            name = cells[0].strip()
            m = re.search(r"(-?[\d.]+)\s*%", cells[1])
            if not name or not m or name.lower().startswith("holding"):
                continue
            weight = float(m.group(1))
            if weight == 0:
                continue
            low = name.lower()
            ticker = next((v for k, v in HIS_NAME_TO_TICKER.items() if k in low), name)
            holdings[ticker] = holdings.get(ticker, 0.0) + weight
        if holdings:
            break
    return holdings


def parse_yieldmax(html: str) -> dict:
    """YieldMax fund pages carry a holdings table headed
    SECURITY NAME | TICKER | CUSIP | SHARES | MARKET VALUE | NET ASSETS | WEIGHTINGS.

    The weight column is "weightings"; cash and option lines often carry no
    ticker and are skipped rather than invented. Negative weights (written
    calls) are dropped, since a short option is not a holding a reader can
    look up.
    """
    soup = BeautifulSoup(html, "lxml")
    holdings = {}
    for table in soup.find_all("table"):
        head = table.find("tr")
        if not head:
            continue
        cols = [c.get_text(" ", strip=True).lower() for c in head.find_all(["th", "td"])]
        if "ticker" not in cols or not any("weight" in c for c in cols):
            continue
        i_t = cols.index("ticker")
        i_w = next(i for i, c in enumerate(cols) if "weight" in c)
        for tr in table.find_all("tr")[1:]:
            cells = [c.get_text(" ", strip=True) for c in tr.find_all("td")]
            if len(cells) <= max(i_t, i_w):
                continue
            ticker = cells[i_t].strip().upper()
            m = re.search(r"(-?[\d.]+)\s*%", cells[i_w])
            if not ticker or ticker in {"-", "\u2014", ""} or not m:
                continue
            # Treasury CUSIPs (9 alphanumerics) and OCC option codes are not
            # holdings a reader can look up, and they crowd out the real exposure.
            if " " in ticker or len(ticker) > 6 or any(ch.isdigit() for ch in ticker):
                continue
            w = float(m.group(1))
            if w <= 0:
                continue
            holdings[ticker] = holdings.get(ticker, 0.0) + w
        if holdings:
            break
    return holdings


def parse_roundhill(html: str) -> dict:
    """Roundhill 0DTE funds hold index options, not equities.

    Their table is Name | Ticker | Weight and the ticker column holds OCC
    option codes like "4SPX  270319C00678800". For these funds that IS the
    position — stripping it the way we strip YieldMax's collateral would leave
    the fund looking empty. So keep the readable Name instead of the raw code,
    which is what a person can actually interpret.
    """
    soup = BeautifulSoup(html, "lxml")
    holdings = {}
    for table in soup.find_all("table"):
        head = table.find("tr")
        if not head:
            continue
        cols = [c.get_text(" ", strip=True).lower() for c in head.find_all(["th", "td"])]
        if "name" not in cols or "weight" not in cols:
            continue
        i_n = cols.index("name")
        i_w = cols.index("weight")
        for tr in table.find_all("tr")[1:]:
            cells = [c.get_text(" ", strip=True) for c in tr.find_all("td")]
            if len(cells) <= max(i_n, i_w):
                continue
            name = cells[i_n].strip()
            m = re.search(r"(-?[\d.]+)\s*%", cells[i_w])
            if not name or not m:
                continue
            w = float(m.group(1))
            if w <= 0:
                continue
            holdings[name] = holdings.get(name, 0.0) + w
        if holdings:
            break
    return holdings


def parse_proshares(html: str) -> dict:
    """ProShares publish a full holdings table headed
    Weight | Ticker | Description | Exposure Value | Market Value | Shares.

    Both a short "Company | Weight" summary and the full table appear on the
    page; prefer the one with a Ticker column so the overlap tool can match
    positions rather than free-text company names.
    """
    soup = BeautifulSoup(html, "lxml")
    holdings = {}
    for table in soup.find_all("table"):
        head = table.find("tr")
        if not head:
            continue
        cols = [c.get_text(" ", strip=True).lower() for c in head.find_all(["th", "td"])]
        if "ticker" not in cols or not any(c.startswith("weight") for c in cols):
            continue
        i_t = cols.index("ticker")
        i_w = next(i for i, c in enumerate(cols) if c.startswith("weight"))
        for tr in table.find_all("tr")[1:]:
            cells = [c.get_text(" ", strip=True) for c in tr.find_all("td")]
            if len(cells) <= max(i_t, i_w):
                continue
            ticker = cells[i_t].strip().upper()
            m = re.search(r"(-?[\d.]+)\s*%", cells[i_w])
            if not ticker or ticker in {"-", "--", "\u2014", ""} or not m:
                continue
            w = float(m.group(1))
            if w <= 0:
                continue
            holdings[ticker] = holdings.get(ticker, 0.0) + w
        if holdings:
            break
    return holdings


def parse_defiance(html: str) -> dict:
    """Defiance keep holdings on a separate /TICKER-full-holdings/ page, headed
    Ticker | Name | CUSIP | ETF Weight | Shares. The fund page itself shows only
    performance, which is why an earlier look found nothing.
    """
    soup = BeautifulSoup(html, "lxml")
    holdings = {}
    for table in soup.find_all("table"):
        head = table.find("tr")
        if not head:
            continue
        cols = [c.get_text(" ", strip=True).lower() for c in head.find_all(["th", "td"])]
        if "ticker" not in cols or not any("weight" in c for c in cols):
            continue
        i_t = cols.index("ticker")
        i_w = next(i for i, c in enumerate(cols) if "weight" in c)
        for tr in table.find_all("tr")[1:]:
            cells = [c.get_text(" ", strip=True) for c in tr.find_all("td")]
            if len(cells) <= max(i_t, i_w):
                continue
            ticker = cells[i_t].strip().upper()
            m = re.search(r"(-?[\d.]+)\s*%", cells[i_w])
            if not ticker or not m:
                continue
            w = float(m.group(1))
            if w <= 0 or " " in ticker or len(ticker) > 6:
                continue
            holdings[ticker] = holdings.get(ticker, 0.0) + w
        if holdings:
            break
    return holdings


def parse_pdf_holdings(html: str) -> dict:
    """Holdings from an issuer's PDF schedule of investments.

    REX and NEOS publish some funds only as quarterly PDFs — no HTML table
    exists anywhere on their sites. This finds the most recent holdings PDF
    linked from the fund page (filenames carry a date and change quarterly, so
    they cannot be hardcoded), downloads it and reads the tables.

    A schedule of investments lists shares and market value but usually not a
    weight column, so weights are computed from each row's value over the total.
    Rows without a usable ticker are skipped rather than guessed at.
    """
    import io

    try:
        import pdfplumber
    except ImportError:
        log.warning("  -> pdfplumber not installed; cannot read PDF holdings")
        return {}

    soup = BeautifulSoup(html, "lxml")
    links = []
    for a in soup.find_all("a", href=True):
        href = a["href"]
        if not href.lower().endswith(".pdf"):
            continue
        label = (a.get_text(" ", strip=True) + " " + href).lower()
        if any(k in label for k in ("holding", "schedule of investment", "soi", "portfolio")):
            if href.startswith("//"):
                href = "https:" + href
            links.append(href)
    if not links:
        log.info("  -> no holdings PDF linked on the page")
        return {}

    # later quarters sort last by the date in the filename
    links.sort()
    url = links[-1]
    log.info("  -> reading holdings PDF %s", url)
    try:
        raw = requests.get(url, headers=ACTIVE_HEADERS, timeout=REQUEST_TIMEOUT).content
    except Exception as exc:  # noqa: BLE001
        log.warning("  -> could not fetch PDF: %s", exc)
        return {}

    rows = []
    try:
        with pdfplumber.open(io.BytesIO(raw)) as pdf:
            for page in pdf.pages[:12]:
                for table in (page.extract_tables() or []):
                    rows.extend(table)
    except Exception as exc:  # noqa: BLE001
        log.warning("  -> could not parse PDF: %s", exc)
        return {}

    values = {}
    for row in rows:
        cells = [(c or "").strip() for c in row]
        if len(cells) < 2:
            continue
        ticker = ""
        for c in cells[:3]:
            t = c.upper().strip()
            if t and t.isalpha() and 1 <= len(t) <= 5:
                ticker = t
                break
        if not ticker:
            continue
        money = None
        for c in reversed(cells):
            m = re.search(r"([\d,]+(?:\.\d+)?)", c.replace("$", ""))
            if m and len(m.group(1).replace(",", "")) >= 4:
                money = float(m.group(1).replace(",", ""))
                break
        if money and money > 0:
            values[ticker] = values.get(ticker, 0.0) + money

    total = sum(values.values())
    if not total:
        log.info("  -> PDF had no readable holdings rows")
        return {}
    return {t: round(v / total * 100, 4) for t, v in sorted(values.items(), key=lambda kv: -kv[1])[:40]}


def parse_listing_only(html: str) -> dict:
    """No holdings — deliberately.

    Some issuers render holdings behind JavaScript or a login, and inventing a
    parser per issuer is a bigger job than it's worth for breadth. Listing a
    fund with an empty holdings dict still gives it a price series, a full
    distribution history from the fallback and therefore a computed yield, so
    it is useful on the site and honest about what's missing. Holdings show as
    unavailable rather than wrong.
    """
    return {}


def parse_bmo(html: str) -> dict:
    """BMO moved to bmogam.com and renders holdings after page load, so this
    needs the browser fetch. The table is headed
    Weight | Name | ISIN | Bloomberg Ticker | ... and the Bloomberg column is
    the clean ticker we want; the name column would need fuzzy matching.
    """
    soup = BeautifulSoup(html, "lxml")
    holdings = {}
    for table in soup.find_all("table"):
        head = table.find("tr")
        if not head:
            continue
        cols = [c.get_text(" ", strip=True).lower() for c in head.find_all(["th", "td"])]
        if "weight" not in cols or not any("bloomberg" in c for c in cols):
            continue
        i_w = cols.index("weight")
        i_t = next(i for i, c in enumerate(cols) if "bloomberg" in c)
        for tr in table.find_all("tr")[1:]:
            cells = [c.get_text(" ", strip=True) for c in tr.find_all("td")]
            if len(cells) <= max(i_w, i_t):
                continue
            ticker = cells[i_t].strip().upper()
            m = re.search(r"(-?[\d.]+)\s*%", cells[i_w])
            if not ticker or ticker in {"—", "-"} or not m:
                continue
            weight = float(m.group(1))
            if weight <= 0:
                continue          # cash and derivative lines can be zero or negative
            holdings[ticker] = holdings.get(ticker, 0.0) + weight
        if holdings:
            break
    return holdings


def _harvest_name_index() -> dict:
    """Map normalised Harvest fund names -> our registry tickers, so that a
    Harvest fund-of-funds row that only gives a fund NAME (its Ticker cell is
    empty) resolves to a real ticker instead of collapsing to 'Harvest'."""
    # Harvest markets some funds under a name that differs from the legal
    # name we hold in the registry, so a pure name match misses them.
    index = {
        "harvest tech leaders income etf": "HTA",
        "harvest canadian equity income leaders etf": "HLIF",
    }
    for f in FUND_REGISTRY:
        if f.issuer == "Harvest ETFs":
            index[_norm_fund_name(f.name)] = f.ticker
    return index


def _norm_fund_name(name: str) -> str:
    import re
    n = name.lower()
    n = re.sub(r"[0-9]+$", "", n.strip())          # strip trailing footnote markers
    n = n.replace("&", "and")
    n = re.sub(r"[^a-z ]", " ", n)
    return re.sub(r"\s+", " ", n).strip()


def parse_harvest(html: str) -> dict:
    """Harvest publishes two different column orders on the same table id:
    equity funds are  Name | Ticker | Weight | Sector | Country
    fund-of-funds are Ticker | ETF Name | Weight | Sector | Country
    and on the fund-of-funds pages the Ticker cell is empty. Reading a fixed
    column index therefore silently produced 'Harvest' for every row (all
    colliding on one dict key). Read the header row instead, and fall back to
    the fund name -> registry ticker when the ticker cell is blank."""
    soup = BeautifulSoup(html, "lxml")
    holdings = {}
    table = soup.select_one('table[id*="_holdings"]')
    if not table:
        raise ValueError("holdings table not found — page structure may have changed")

    rows = table.select("tr")
    if not rows:
        raise ValueError("holdings table has no rows")

    header = [c.get_text(strip=True).lower() for c in rows[0].select("th,td")]
    def col(*names, default=None):
        for i, h in enumerate(header):
            for n in names:
                if n in h:
                    return i
        return default

    i_ticker = col("ticker", default=0)
    i_name = col("name", default=1)
    i_weight = col("weight", "%", default=2)
    name_index = _harvest_name_index()

    for row in rows[1:]:
        cells = [c.get_text(strip=True) for c in row.select("td")]
        if len(cells) <= max(i_ticker, i_weight):
            continue

        raw_ticker = cells[i_ticker] if i_ticker < len(cells) else ""
        raw_name = cells[i_name] if i_name is not None and i_name < len(cells) else ""
        weight = cells[i_weight]

        # "REGN US" -> "REGN"; blank ticker -> resolve the fund name
        ticker = raw_ticker.split()[0] if raw_ticker else ""
        if not ticker and raw_name:
            ticker = name_index.get(_norm_fund_name(raw_name), raw_name.strip())
        if not ticker or not weight:
            continue

        try:
            w = float(weight.replace("%", "").replace(",", ""))
        except ValueError:
            continue
        if w <= 0:
            continue
        # same underlying can appear twice (e.g. two share classes) — add, don't clobber
        holdings[ticker] = round(holdings.get(ticker, 0.0) + w, 4)

    return holdings


def _evolve_from_page(html: str) -> dict:
    """Read the holdings table printed on the fund page itself.

    Columns are Name, Weight, Ticker, Sector. The ticker is Bloomberg style
    ("BNS CN EQUITY"), so only the leading symbol is kept.
    """
    import re
    from bs4 import BeautifulSoup
    soup = BeautifulSoup(html, "lxml")
    for table in soup.find_all("table"):
        head = [th.get_text(strip=True).lower() for th in table.find_all("th")]
        if "weight" not in head or "ticker" not in head:
            continue
        wi, ti = head.index("weight"), head.index("ticker")
        out = {}
        for tr in table.find_all("tr"):
            cells = [td.get_text(strip=True) for td in tr.find_all("td")]
            if len(cells) <= max(wi, ti):
                continue
            sym = cells[ti].split()[0] if cells[ti] else ""
            m = re.search(r"(\d+(?:\.\d+)?)", cells[wi].replace(",", ""))
            if not sym or not m:
                continue
            w = float(m.group(1))
            if w > 0:
                out[sym] = round(w, 2)
        if out:
            return out
    return {}

def _globalx_ca_pairs(html: str) -> dict:
    """Read Global X Canada holdings from the rendered fund page.

    The page shows "Top Holdings" (what the fund owns directly, often just
    another Global X fund) and "Top Underlying Holdings" (the companies that
    sit inside it). The underlying list is the useful one; fall back to the
    direct list when a fund holds securities itself.
    """
    import re
    from bs4 import BeautifulSoup
    soup = BeautifulSoup(html, "lxml")
    text_nodes = [t.strip() for t in soup.stripped_strings]

    def block(title):
        try:
            start = text_nodes.index(title)
        except ValueError:
            return {}
        out = {}
        i = start + 1
        while i < len(text_nodes) - 1:
            name = text_nodes[i]
            nxt = text_nodes[i + 1]
            if name in ("Top Holdings", "Top Underlying Holdings", "Documents"):
                break
            m = re.fullmatch(r"(-?\d+(?:\.\d+)?)%", nxt)
            if m and name not in ("Weight", "Security Name"):
                w = float(m.group(1))
                if w > 0 and not re.match(r"(?i)cash|as at", name):
                    out[name] = round(w, 2)
                i += 2
                continue
            i += 1
        return out

    under = block("Top Underlying Holdings")
    if under:
        return under
    return block("Top Holdings")

def _rex_pairs(html: str) -> dict:
    """Read REX holdings from the grid printed on the fund page."""
    import re
    from bs4 import BeautifulSoup
    soup = BeautifulSoup(html, "lxml")
    cells = [t.strip() for t in soup.stripped_strings]
    try:
        start = cells.index("Weighting")
    except ValueError:
        return {}
    out = {}
    sym_re = re.compile(r"^[A-Z][A-Z.\-]{0,5}$")
    pct_re = re.compile(r"^(-?\d+(?:\.\d+)?)%$")
    i = start + 1
    while i < len(cells) - 1:
        sym = cells[i]
        if sym_re.match(sym):
            for j in range(i + 1, min(i + 5, len(cells))):
                m = pct_re.match(cells[j])
                if m:
                    w = float(m.group(1))
                    if w > 0 and sym not in ("USD", "CASH"):
                        out[sym] = round(w, 2)
                    i = j
                    break
        i += 1
    return out


def parse_rex(html: str) -> dict:
    """The PDF route first, then the grid on the page."""
    try:
        rows = parse_pdf_holdings(html)
        if rows:
            return rows
    except Exception:  # noqa: BLE001
        pass
    return _rex_pairs(html)

def parse_evolve(html: str) -> dict:
    import csv
    import re

    match = re.search(r"https://evolveetfs\.com/wp-content/uploads/holdings/[\w\-]+\.csv(?:\?[^\s\"']*)?", html)
    if not match:
        raise ValueError("could not locate the holdings CSV link on the fund page")
    csv_url = match.group(0)

    try:
        csv_resp = requests.get(csv_url, headers=REQUEST_HEADERS, timeout=REQUEST_TIMEOUT)
        csv_resp.raise_for_status()
    except Exception as exc:  # noqa: BLE001
        page_rows = _evolve_from_page(html)
        if page_rows:
            log.info("  -> CSV refused (%s); used the table on the page", exc)
            return page_rows
        raise

    reader = csv.DictReader(csv_resp.text.splitlines())
    holdings = {}
    ticker_re = re.compile(r"^[A-Z]{1,6}([./][A-Z]{1,3})?$")
    for row in reader:
        ticker_raw = (row.get("TICKER") or "").strip()
        name = (row.get("SECURITY_NAME") or "").strip().lower()
        weight_raw = (row.get("PORTFOLIO_MWEIGHT") or "").strip()
        if not ticker_raw or not weight_raw or "option" in name:
            continue
        ticker = ticker_raw.split()[0]
        if not ticker_re.match(ticker):
            continue
        try:
            w = float(weight_raw) * 100
        except ValueError:
            continue
        if w <= 0:
            continue
        holdings[ticker] = round(w, 2)

    if not holdings:
        holdings = _evolve_from_page(html)
    if not holdings:
        raise ValueError(f"CSV fetched from {csv_url} but no rows parsed")
    return holdings




def parse_globalx_us(html: str) -> dict:
    """Confirmed CSV schema (Aug 2026): % of Net Assets,Ticker,Name,SEDOL,
    Market Price ($), Shares Held, Market Value ($), with a title line before
    the header row. CSV lives at assets.globalxetfs.com/funds/holdings/
    {ticker}_full-holdings_{YYYYMMDD}.csv, linked from the fund page."""
    import csv
    import re

    soup = BeautifulSoup(html, "lxml")
    csv_link_tag = soup.select_one("a[href*='/funds/holdings/'][href$='.csv']")
    if not csv_link_tag:
        match = re.search(r"https://assets\.globalxetfs\.com/funds/holdings/[\w\-]+\.csv", html)
        if not match:
            raise ValueError("could not locate the holdings CSV link on the fund page")
        csv_url = match.group(0)
    else:
        csv_url = csv_link_tag.get("href")

    csv_resp = requests.get(csv_url, headers=REQUEST_HEADERS, timeout=REQUEST_TIMEOUT)
    csv_resp.raise_for_status()

    lines = csv_resp.text.splitlines()
    header_idx = 0
    for i, line in enumerate(lines):
        if line.strip().lower().startswith(("% of net assets", "\ufeff% of net assets")):
            header_idx = i
            break

    reader = csv.DictReader(lines[header_idx:])
    holdings = {}
    ticker_re = re.compile(r"^[A-Z]{1,6}([./][A-Z]{1,3})?$")
    skip_names = ("cash", "payable", "receivable", "future", "index", "swap")
    for row in reader:
        ticker = (row.get("Ticker") or "").strip()
        name = (row.get("Name") or "").strip().lower()
        weight_raw = (row.get("% of Net Assets") or row.get("\ufeff% of Net Assets") or "").strip()
        if not ticker or not weight_raw or not ticker_re.match(ticker):
            continue
        if any(bad in name for bad in skip_names):
            continue
        try:
            w = float(weight_raw.replace("%", "").replace(",", ""))
        except ValueError:
            continue
        if w <= 0:
            continue
        holdings[ticker] = w
      
    if not holdings:
        raise ValueError(f"CSV fetched from {csv_url} but no rows parsed")
    return holdings

def parse_jpmorgan_xls(html: str) -> dict:
    """JPMorgan publishes a 'Download all holdings (XLS)' export as a real .xlsx
    file behind a static, parameterized URL — no JS rendering needed. This
    function is called with the RAW BYTES of that xlsx (see fetch_binary in
    run()), passed through as a latin-1 decoded string so it can flow through
    the same str-based parser interface as every other issuer."""
    import io
    import re
    from openpyxl import load_workbook
    raw_bytes = html.encode("latin-1")
    wb = load_workbook(io.BytesIO(raw_bytes), data_only=True, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header_idx = None
    sym_col = pct_col = None
    for i, row in enumerate(rows):
        cells = [str(c).strip().lower() if c is not None else "" for c in row]
        if "ticker" in cells and any("net assets" in c for c in cells):
            header_idx = i
            sym_col = cells.index("ticker")
            pct_col = next(j for j, c in enumerate(cells) if "net assets" in c)
            break
    if header_idx is None:
        raise ValueError("could not find a 'Symbol' + '% of Net Assets' header row in the xlsx")
    holdings = {}
    ticker_re = re.compile(r"^[A-Z]{1,6}([./][A-Z]{1,3})?$")
    for row in rows[header_idx + 1:]:
        if row is None or len(row) <= max(sym_col, pct_col):
            continue
        ticker = str(row[sym_col]).strip() if row[sym_col] is not None else ""
        pct_raw = row[pct_col]
        if not ticker or not ticker_re.match(ticker) or pct_raw is None:
            continue
        try:
            pct_str = str(pct_raw).replace("%", "").strip()
            w_val = float(pct_str)
            w = w_val
        except (TypeError, ValueError):
            continue
        if w <= 0:
            continue
        holdings[ticker] = round(w, 2)
    if not holdings:
        raise ValueError("xlsx parsed but no holdings rows matched")
    return holdings

def parse_neos_csv(html: str) -> dict:
    """NEOS publishes a plain-text CSV export at a static admin-ajax URL
    (columns: Date, Account, StockTicker, Cusip, SecurityName, Shares, Price,
    MarketValue, Weightings, ...). No JS rendering needed."""
    import csv
    import re
    reader = csv.DictReader(html.splitlines())
    holdings = {}
    ticker_re = re.compile(r"^[A-Z]{1,6}([./][A-Z]{1,3})?$")
    for row in reader:
        ticker = (row.get("StockTicker") or "").strip()
        weight_raw = (row.get("Weightings") or "").strip()
        if not ticker or not weight_raw or not ticker_re.match(ticker):
            continue
        try:
            w = float(weight_raw.replace("%", "").replace(",", ""))
        except ValueError:
            continue
        if w <= 0:
            continue
        holdings[ticker] = round(w, 2)
    if not holdings:
        raise ValueError("NEOS csv fetched but no holdings rows parsed")
    return holdings

def parse_amplify_rendered(html: str) -> dict:
    """Amplify's holdings pages populate a table client-side from Firestore —
    this parser receives already-RENDERED html (see fetch_rendered in run())
    with the table already populated. Table columns: Name | Ticker |
    Market Value (%) | CUSIP | Shares | Market Value ($)."""
    import re
    soup = BeautifulSoup(html, "lxml")
    table = soup.select_one("table")
    if not table:
        raise ValueError("holdings table not found in rendered page")
    rows = table.select("tr")
    holdings = {}
    ticker_re = re.compile(r"^[A-Z]{1,6}([./][A-Z]{1,3})?$")
    for row in rows[1:]:
        cells = [c.get_text(strip=True) for c in row.select("td")]
        if len(cells) < 3:
            continue
        ticker, weight = cells[1], cells[2]
        if not ticker or not weight or not ticker_re.match(ticker):
            continue
        try:
            w = float(weight.replace("%", "").replace(",", ""))
        except ValueError:
            continue
        if w <= 0:
            continue
        holdings[ticker] = round(w, 2)
    if not holdings:
        raise ValueError("rendered page fetched but no holdings rows parsed")
    return holdings

def parse_globalx_ca_rendered(html: str) -> dict:
    """Global X Canada's Holdings tab is client-rendered and, worse, its
    'Top Underlying Holdings' section lists company NAMES, not tickers —
    unusable for ticker-based overlap without a name->ticker map. Instead we
    read the 'Top Holdings' section just above it, which lists the actual
    wrapped fund/ticker this ETF holds, e.g. 'GLOBAL X S&P/TSX 60 INDEX ETF
    (CNDX)  100.00%'. This gives one real ticker per fund rather than
    dozens of unmatched names — a genuine, if partial, overlap signal."""
    import re
    match = re.search(r"\(([A-Z]{1,6})\)[\s\S]{1,500}?(\d+\.\d+)%", html)
    if not match:
        raise ValueError("could not find a wrapped-ticker top holding in rendered page")
    ticker, weight = match.group(1), match.group(2)
    w = float(weight)
    if w <= 0:
        raise ValueError("parsed a zero-weight top holding")
    return {ticker: round(w, 2)}
  





_globalx_ca_original = parse_globalx_ca_rendered


def parse_globalx_ca_rendered(html: str) -> dict:  # noqa: F811
    """Table first, then the div layout the site moved to."""
    try:
        rows = _globalx_ca_original(html)
        if rows:
            return rows
    except Exception:  # noqa: BLE001
        pass
    return _globalx_ca_pairs(html)

# ---------------------------------------------------------------------------
# 2b. FUND STATS (AUM / NAV / yield / MER)
#
# None of this lives in the holdings files — it sits in the "key facts" block
# on each issuer's fund page. Every issuer renders that block differently in
# HTML, but they all render it as label-then-value in READING ORDER, so rather
# than nine bespoke selectors we flatten the page to text lines once and look
# for a label, then scan the next few lines for the first value that matches
# the shape we expect (money / percent). That survives markup changes.
#
# Issuers with no AUM on the fetched page (Hamilton, J.P. Morgan) simply
# return {} — the site treats missing stats as "not collected", not as zero.
# ---------------------------------------------------------------------------

MONEY_RE = re.compile(
    r"\$\s?([\d,]+(?:\.\d+)?)\s*(billion|million|thousand|bn|mm|[BMK])?\b",
    re.IGNORECASE,
)
PERCENT_RE = re.compile(r"(-?[\d.]+)\s*%")

MULTIPLIER = {
    "b": 1000.0, "bn": 1000.0, "billion": 1000.0,
    "m": 1.0, "mm": 1.0, "million": 1.0,
    "k": 0.001, "thousand": 0.001,
}


def _text_lines(html: str) -> list:
    soup = BeautifulSoup(html, "lxml")
    for tag in soup(["script", "style", "noscript"]):
        tag.decompose()
    raw = soup.get_text("\n")
    return [ln.strip() for ln in raw.split("\n") if ln.strip()]


def _money_to_millions(text: str):
    """'$8.36 billion' / '$1.275 B' / '$2133.67M' / '$194,678,118' -> millions."""
    m = MONEY_RE.search(text)
    if not m:
        return None
    value = float(m.group(1).replace(",", ""))
    unit = (m.group(2) or "").lower()
    if unit:
        return round(value * MULTIPLIER.get(unit, 1.0), 3)
    # a bare figure is in dollars, not millions
    return round(value / 1_000_000, 3)


def _find_after(lines: list, labels: list, pattern, lookahead: int = 6):
    """Find a label line, then return the first match for `pattern` in the rest
    of that line or in the next few lines. Tooltips and blank cells often sit
    between a label and its value, which is why we look ahead rather than
    trusting the very next line."""
    for i, line in enumerate(lines):
        low = line.lower().strip().lstrip("\u2022 ").rstrip("*: ").strip()
        matched = None
        for w in labels:
            if low == w or re.match(re.escape(w) + r"\b", low):
                matched = w
                break
        if not matched:
            continue
        rest = line.lower().split(matched, 1)[-1]
        m = pattern.search(line[len(line) - len(rest):]) if rest else None
        if m:
            return matched, m.group(0).strip()
        for candidate in lines[i + 1 : i + 1 + lookahead]:
            m = pattern.search(candidate)
            if m:
                return matched, m.group(0).strip()
    return None


AUM_LABELS = ["net aum", "aum", "net assets", "fund total net assets",
              "total net assets", "fund size", "assets under management",
              # Hamilton's fund-facts table just says "Assets" ($2489.4M CAD)
              "assets",
              # Harvest High Income Shares print "AUM*" with a footnote marker
              "aum*"]
NAV_LABELS = ["nav", "net asset value", "closing nav"]
YIELD_LABELS = ["current annualized yield", "current yield", "distribution yield",
                "trailing 12-month yield", "yield %", "annualized distribution yield",
                "30-day sec yield", "annualized yield"]
MER_LABELS = ["management expense ratio", "mer", "total expense ratio",
              "total annual fund operating expenses", "expense ratio"]
FEE_LABELS = ["management fee", "mgmt fee"]


def parse_fund_stats(html: str) -> dict:
    """Issuer-agnostic key-facts reader. Returns only the keys it actually
    found, so a partially-populated page degrades to partial stats.

    Yield and MER are stored with the label the issuer used, because those
    labels are not interchangeable across issuers: Global X US publishes a
    30-day SEC yield (0.02% for QYLD) while Global X Canada publishes an
    annualised distribution yield (6.65% for CNCC). Showing either as a bare
    "Yield" would be actively misleading, so the site prints the issuer's own
    wording.
    """
    lines = _text_lines(html)
    stats = {}

    hit = _find_after(lines, AUM_LABELS, MONEY_RE)
    if hit:
        millions = _money_to_millions(hit[1])
        # sanity floor/ceiling: $0.0M or $50T means we matched the wrong line
        if millions is not None and 0.05 <= millions <= 5_000_000:
            stats["aum_musd"] = millions
            stats["aum_display"] = hit[1]

    hit = _find_after(lines, NAV_LABELS, MONEY_RE)
    if hit:
        stats["nav"] = hit[1]

    hit = _find_after(lines, YIELD_LABELS, PERCENT_RE)
    if hit:
        stats["yield"] = hit[1]
        stats["yield_label"] = hit[0]

    hit = _find_after(lines, MER_LABELS, PERCENT_RE)
    if hit:
        stats["mer"] = hit[1]
        stats["mer_label"] = hit[0]

    # Purpose builds its pages without <table> elements, so there is no history
    # to parse — but the key-facts block does carry the most recent payment, and
    # one real number beats an empty distributions section.
    hit = _find_after(lines, ["last distribution", "latest distribution"], MONEY_RE)
    if hit:
        stats["last_amount"] = float(hit[1].replace("$", "").replace(",", "").strip())
        for ln in lines:
            m = re.search(r"\((20\d{2}-\d{2}-\d{2})\)", ln)
            if m:
                stats["last_ex_date"] = m.group(1)
                break

    hit = _find_after(lines, FEE_LABELS, PERCENT_RE)
    if hit:
        stats["mgmt_fee"] = hit[1]

    return stats


# Which issuers expose key facts on a page we can reach, and where.
# Value is either None (reuse the holdings page we already fetched) or a
# callable turning a Fund into the profile URL to fetch separately.
STATS_SOURCES = {
    "harvest": None,
    # PDF-only issuers still show AUM on the fund page
    "pdf_holdings": None,
    # ProShares pages carry Net Assets beside the holdings table
    "proshares": None,
    # Roundhill pages carry AUM beside the holdings table
    "roundhill": None,
    # YieldMax pages carry Net Assets beside the holdings table
    "yieldmax": None,
    # High Income Shares carry AUM*, NAV and yield on the same page as holdings
    "harvest_his": None,
    # BMO's rendered page carries "Net assets (M)" alongside the holdings table
    "bmo": None,
    "purpose_single": None,
    "evolve": None,
    "globalx_us": None,
    "globalx_ca_rendered": None,
    "neos_csv": lambda f: f"https://neosfunds.com/{f.ticker.lower()}/",
    "amplify_rendered": lambda f: f"https://amplifyetfs.com/{f.ticker.lower()}/",
    # hamilton: fund page carries NAV + yield but no AUM
    "hamilton": None,
    # jpmorgan: holdings come from an .xlsx export, key facts are behind JS
}


PARSERS: dict[str, Callable[[str], dict]] = {
    "rex": parse_rex,
    "pdf_holdings": parse_pdf_holdings,
    "defiance": parse_defiance,
    "proshares": parse_proshares,
    "roundhill": parse_roundhill,
    "yieldmax": parse_yieldmax,
    "harvest_his": parse_harvest_his,
    "listing_only": parse_listing_only,
    "hamilton": parse_hamilton,
    "bmo": parse_bmo,
    "harvest": parse_harvest,
    "evolve": parse_evolve,
    "globalx_us": parse_globalx_us,
    "jpmorgan_xls": parse_jpmorgan_xls,
    "neos_csv": parse_neos_csv,
    "amplify_rendered": parse_amplify_rendered,
    "globalx_ca_rendered": parse_globalx_ca_rendered,
    "purpose_single": parse_purpose_single,
}


# ---------------------------------------------------------------------------
# 3. FETCH + ORCHESTRATION
# ---------------------------------------------------------------------------

def fetch(url: str) -> str:
    """One retry with the other header set on a refusal. Issuers disagree about
    which requests they like, and the disagreement shows up as 403/406/429, so
    trying the other identity once is cheaper than maintaining a perfect map."""
    resp = requests.get(url, headers=ACTIVE_HEADERS, timeout=REQUEST_TIMEOUT)
    if resp.status_code in (401, 403, 406, 429, 503):
        other = LEGACY_HEADERS if ACTIVE_HEADERS is REQUEST_HEADERS else REQUEST_HEADERS
        log.info("  -> %s on %s, retrying with the other header set",
                 resp.status_code, url)
        time.sleep(1.0)
        resp = requests.get(url, headers=other, timeout=REQUEST_TIMEOUT)
    resp.raise_for_status()
    return resp.text


def fetch_binary(url: str) -> str:
    """For endpoints that return a binary file (e.g. JPMorgan's .xlsx export).
    Returns the raw bytes decoded as latin-1 (a lossless 1:1 byte mapping),
    so the bytes can be re-encoded exactly by the parser and every parser
    can keep the same str-in/dict-out signature."""
    resp = requests.get(url, headers=ACTIVE_HEADERS, timeout=REQUEST_TIMEOUT)
    resp.raise_for_status()
    return resp.content.decode("latin-1")

def fetch_rendered(url: str, wait_selector: str = None, wait_ms: int = 25000,
                   click_selector: str = None) -> str:
    """For pages that only populate their holdings table via client-side JS
    (Amplify's Firestore-backed holdings pages, Global X Canada's Holdings
    tab). Uses a headless Chromium via Playwright, waits for either a given
    CSS selector to appear or a flat delay, then returns the fully-rendered
    page HTML, which the matching parser then reads with BeautifulSoup
    exactly like a normal static fetch."""
    from playwright.sync_api import sync_playwright
    with sync_playwright() as p:
        # BMO's server aborts the HTTP/2 handshake with headless Chromium
        # (ERR_HTTP2_PROTOCOL_ERROR); forcing HTTP/1.1 makes the page load and
        # costs nothing on sites that were already fine.
        browser = p.chromium.launch(args=["--disable-http2"])
        page = browser.new_page(user_agent=ACTIVE_HEADERS["User-Agent"])
        try:
            # BMO in particular is slow to first byte; a rendered page is worth
            # waiting longer for than a plain fetch.
            page.goto(url, timeout=90000, wait_until="domcontentloaded")
            if click_selector:
                clicked = False
                try:
                    el = page.query_selector(click_selector)
                    if el:
                        el.scroll_into_view_if_needed(timeout=5000)
                        page.wait_for_timeout(400)
                        el.click(timeout=8000)
                        clicked = True
                except Exception as exc:  # noqa: BLE001
                    log.info("  -> normal click failed (%s)", exc)
                if not clicked:
                    try:
                        page.eval_on_selector(click_selector, "el => el.click()")
                        clicked = True
                    except Exception as exc:  # noqa: BLE001
                        log.warning("  -> could not click %s (%s)", click_selector, exc)
                if clicked:
                    page.wait_for_timeout(4000)
            if wait_selector:
                try:
                    page.wait_for_selector(wait_selector, timeout=wait_ms)
                except Exception:
                    # the table may be named differently on this page; take
                    # what rendered rather than returning nothing at all
                    log.warning("  -> %s did not appear, reading page as-is", wait_selector)
                    page.wait_for_timeout(6000)
            else:
                page.wait_for_timeout(min(wait_ms, 8000))
            return page.content()
        finally:
            browser.close()




def describe_response(html) -> str:
    """When a parser can't find what it expects, the useful question is what the
    server actually sent. A bot-challenge page, a consent wall and a genuinely
    restructured page all look identical to a selector that returns None, so log
    enough of the body to tell them apart on the next run."""
    if html is None:
        return "no response body — the fetch itself failed"
    try:
        soup = BeautifulSoup(html, "lxml")
        title = soup.title.get_text(strip=True) if soup.title else "(no <title>)"
        tables = len(soup.find_all("table"))
        table_ids = [t.get("id") or t.get("class") or "(unnamed)"
                     for t in soup.find_all("table")[:5]]
        text = " ".join(soup.get_text(" ").split())[:180]
        return (f"body was {len(html)} chars, title={title!r}, {tables} table(s) "
                f"{table_ids}, text starts: {text!r}")
    except Exception:  # noqa: BLE001 — diagnostics must never mask the real error
        return f"body was {len(html)} chars, could not be parsed for diagnostics"


DATE_RE = re.compile(r"(20\d{2})[/\-](\d{1,2})[/\-](\d{1,2})|(\d{1,2})/(\d{1,2})/(20\d{2})")
AMOUNT_RE = re.compile(r"\$\s?([\d,]+\.\d{2,6})")


def _iso_date(text: str):
    m = DATE_RE.search(text or "")
    if not m:
        return None
    if m.group(1):
        y, mo, d = m.group(1), m.group(2), m.group(3)
    else:
        mo, d, y = m.group(4), m.group(5), m.group(6)
    try:
        return "%04d-%02d-%02d" % (int(y), int(mo), int(d))
    except ValueError:
        return None


def parse_distributions(html: str) -> list:
    """Pull distribution history out of whatever table the issuer put it in.

    Issuers vary wildly: Harvest splits history across a dozen per-year tables
    (tablepress-hhl_distributions, -no-2, -no-3 ...), Hamilton uses a single
    table, and column order differs everywhere. So rather than per-issuer
    selectors, find any table whose header mentions a distribution-ish date and
    an amount, then read columns by header name.

    Returns newest-first [{ex_date, pay_date, amount, frequency}], de-duplicated
    on ex_date, capped at 60 rows (five years of monthly payers) so a fund with
    a decade of history doesn't bloat funds.json.
    """
    soup = BeautifulSoup(html, "lxml")
    rows_out = {}

    for table in soup.find_all("table"):
        head_cells = table.find("tr")
        if not head_cells:
            continue
        headers = [c.get_text(" ", strip=True).lower()
                   for c in head_cells.find_all(["th", "td"])]
        joined = " ".join(headers)
        if not any(k in joined for k in ("ex-dividend", "ex dividend", "ex-div",
                                         "ex-distribution", "ex date")):
            continue
        if not any(k in joined for k in ("amount", "cash", "distribution", "class a", "per unit", "$")):
            continue

        def col(*names):
            for i, h in enumerate(headers):
                if any(nm in h for nm in names):
                    return i
            return None

        i_ex = col("ex-dividend", "ex dividend", "ex-div", "ex-distribution", "ex date")
        i_pay = col("payment date", "pay date", "payable")
        i_amt = col("class a", "amount", "per unit", "cash distribution", "amount ($)")
        i_freq = col("frequency")

        for tr in table.find_all("tr")[1:]:
            cells = [c.get_text(" ", strip=True) for c in tr.find_all("td")]
            if not cells or i_ex is None or i_ex >= len(cells):
                continue
            ex = _iso_date(cells[i_ex])
            if not ex:
                continue
            amt = None
            if i_amt is not None and i_amt < len(cells):
                m = AMOUNT_RE.search(cells[i_amt])
                if m:
                    amt = float(m.group(1).replace(",", ""))
            if amt is None:  # fall back to the first money-looking cell
                for c in cells:
                    m = AMOUNT_RE.search(c)
                    if m:
                        amt = float(m.group(1).replace(",", ""))
                        break
            if amt is None:
                continue
            rec = {"ex_date": ex, "amount": round(amt, 6)}
            pay = _iso_date(cells[i_pay]) if (i_pay is not None and i_pay < len(cells)) else None
            if pay:
                rec["pay_date"] = pay
            if i_freq is not None and i_freq < len(cells) and cells[i_freq]:
                rec["frequency"] = cells[i_freq]
            rows_out.setdefault(ex, rec)   # first table wins on duplicates

    out = sorted(rows_out.values(), key=lambda r: r["ex_date"], reverse=True)
    return out[:60]


def distribution_summary(dists: list) -> dict:
    """Trailing-12-month total and the latest payment — the two numbers an
    income investor actually looks at. Deliberately NOT a yield: that needs a
    price, and dividing by NAV would silently produce a different figure from
    the issuer's own published yield."""
    if not dists:
        return {}
    latest = dists[0]
    cutoff = (datetime.now(timezone.utc).date() - timedelta(days=365)).isoformat()
    ttm = [d["amount"] for d in dists if d["ex_date"] >= cutoff]
    out = {
        "last_amount": latest["amount"],
        "last_ex_date": latest["ex_date"],
        "count": len(dists),
    }
    if latest.get("frequency"):
        out["frequency"] = latest["frequency"]
    if ttm:
        out["ttm_total"] = round(sum(ttm), 4)
        out["ttm_payments"] = len(ttm)
    return out


SEED: dict = {}
DH_BASE = "https://dividendhistory.org/payout/"
DH_ATTRIBUTION = "dividendhistory.org"


def dividendhistory_urls(fund: Fund) -> list:
    """Most Canadian listings sit under /payout/tsx/, but not all — TECY is a
    Canadian fund filed under the bare path, and guessing from region alone
    silently loses it. Try the likely shape first, then the other."""
    t = fund.ticker.upper()
    tsx = f"{DH_BASE}tsx/{t}/"
    plain = f"{DH_BASE}{t}/"
    return [tsx, plain] if fund.region == "CAD" else [plain, tsx]


def parse_dividendhistory(html: str) -> list:
    """Distribution history from dividendhistory.org.

    Their table leads with FUTURE rows flagged 'unconfirmed/estimated' — for
    HMAX that was two projected payments dated after today. Ingesting those
    would inflate trailing-twelve-month totals and draw distributions on the
    cadence chart that were never paid, so anything unconfirmed or dated in the
    future is dropped. That rule is deliberate; don't relax it to gain rows.
    """
    soup = BeautifulSoup(html, "lxml")
    table = soup.find("table", id="dividend-table") or soup.find("table")
    if not table:
        return []

    today = datetime.now(timezone.utc).date().isoformat()
    out = []
    for tr in table.find_all("tr")[1:]:
        cells = [c.get_text(" ", strip=True) for c in tr.find_all("td")]
        if len(cells) < 3:
            continue
        ex = _iso_date(cells[0])
        if not ex or ex > today:
            continue                                   # not paid yet
        status = cells[3].lower() if len(cells) > 3 else ""
        if "unconfirm" in status or "estimat" in status:
            continue                                   # a projection, not a payment
        m = AMOUNT_RE.search(cells[2])
        if not m:
            continue
        rec = {"ex_date": ex,
               "amount": round(float(m.group(1).replace(",", "")), 6),
               "source": DH_ATTRIBUTION}
        pay = _iso_date(cells[1])
        if pay:
            rec["pay_date"] = pay
        out.append(rec)

    out.sort(key=lambda r: r["ex_date"], reverse=True)
    return out[:60]


def fetch_dividendhistory(fund: Fund) -> list:
    """Fallback only. Issuer-published history always wins, because it is the
    primary record; this fills funds whose issuer publishes nothing parseable."""
    for url in dividendhistory_urls(fund):
        try:
            log.info("  -> distributions fallback %s", url)
            html = fetch(url)
            time.sleep(DELAY_BETWEEN_REQUESTS)
            rows = parse_dividendhistory(html)
            if rows:
                log.info("  -> %d distributions from %s, latest %s",
                         len(rows), DH_ATTRIBUTION, rows[0]["ex_date"])
                return rows
            log.info("  -> no usable rows at %s", url)
        except Exception as exc:  # noqa: BLE001
            log.info("  -> %s: %s", url, exc)
    log.warning("  -> no distributions found anywhere for %s", fund.ticker)
    return []


def profile_html(fund: Fund, holdings_html: str) -> str:
    """The page carrying key facts AND distribution history.

    For most issuers that's the same page we just fetched for holdings. NEOS and
    Amplify are the exceptions: their holdings arrive as a CSV / separate URL, so
    their fund page has to be fetched once — and it is that page, not the CSV,
    that holds the distribution tables. Fetch it once and reuse for both.
    """
    if fund.parser not in STATS_SOURCES:
        return holdings_html
    source = STATS_SOURCES[fund.parser]
    if source is None:
        return holdings_html
    url = source(fund)
    log.info("  -> profile page %s", url)
    html = fetch(url)
    time.sleep(DELAY_BETWEEN_REQUESTS)
    return html


def collect_stats(fund: Fund, page_html: str) -> dict:
    """Never fatal: a stats failure leaves holdings intact and logs a warning,
    because a fund with holdings and no AUM is still useful."""
    if fund.parser not in STATS_SOURCES:
        return {}
    try:
        stats = parse_fund_stats(page_html)
        if not stats:
            log.info("  -> no key facts found for %s", fund.ticker)
        return stats
    except Exception as exc:  # noqa: BLE001
        log.warning("  -> stats FAILED for %s: %s", fund.ticker, exc)
        return {}


SEED_PATH = Path("data/seed_holdings.json")


def load_seed() -> dict:
    """Holdings captured by hand from an issuer's own page.

    Some issuers (BMO) answer a residential browser but hang forever on the
    GitHub Actions IP range, so no timeout or header change reaches them. The
    numbers here came from their published holdings table; the fund page shows
    the capture date and says where it came from, rather than pretending this
    was scraped today.
    """
    if not SEED_PATH.exists():
        return {}
    try:
        return json.loads(SEED_PATH.read_text())
    except Exception as exc:  # noqa: BLE001
        log.warning("could not read %s: %s", SEED_PATH, exc)
        return {}


def load_previous(path: Path = OUTPUT_PATH) -> dict:
    """The previous run's output, used to carry a fund forward when today's
    fetch fails. An issuer blocking us for a day is not the same event as a
    fund ceasing to exist, and the site shouldn't render them identically."""
    if not path.exists():
        return {}
    try:
        return json.loads(path.read_text())
    except Exception as exc:  # noqa: BLE001
        log.warning("Could not read previous %s: %s", path, exc)
        return {}


def carry_forward(fund: Fund, previous: dict) -> None:
    """Keep the last good holdings/stats and mark them stale, so the fund page
    shows real numbers with an honest 'as of' date instead of going blank."""
    prev = previous.get(fund.ticker)
    if not prev or not prev.get("holdings"):
        return
    fund.holdings = prev.get("holdings", {})
    fund.stats = prev.get("stats", {}) or {}
    fund.distributions = prev.get("distributions", []) or []
    fund.as_of = prev.get("as_of", "") or prev.get("last_seen", "")
    fund.stale = True
    log.info("  -> carrying forward %d holdings from %s",
             len(fund.holdings), fund.as_of or "an earlier run")


def run(registry: list[Fund]) -> list[Fund]:
    previous = load_previous()
    global SEED
    SEED = load_seed()
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    for fund in registry:
        log.info("Fetching %s (%s) from %s", fund.ticker, fund.issuer, fund.holdings_url)
        html = None
        global ACTIVE_HEADERS
        ACTIVE_HEADERS = HEADERS_BY_PARSER.get(fund.parser, REQUEST_HEADERS)
        try:
            if fund.parser == "jpmorgan_xls":
                html = fetch_binary(fund.holdings_url)
            elif fund.needs_browser:
                html = fetch_rendered(fund.holdings_url, fund.wait_selector,
                                      click_selector=fund.click_selector)
            else:
                html = fetch(fund.holdings_url)
            parser = PARSERS[fund.parser]
            fund.holdings = parser(html)
            fund.fetched_ok = bool(fund.holdings)
            if not fund.fetched_ok:
                log.warning("  -> no holdings parsed; keeping the rest of the page")

            # stats do not depend on holdings: read them either way
            fund.as_of = today
            fund.stale = False
            page = profile_html(fund, html)
            fresh = collect_stats(fund, page)
            prior = (previous.get(fund.ticker) or {}).get("stats") or {}
            merged = dict(prior)
            for _k, _v in fresh.items():
                if _v is not None and _v != "":
                    merged[_k] = _v  # a fresh reading always wins
            # anything the page did not give up this time keeps the older
            # figure rather than going blank on the site
            fund.stats = merged
            if prior and not fresh:
                log.warning("  -> no stats parsed; kept the previous figures")
            if True:
                try:
                    fund.distributions = parse_distributions(page)
                    if not fund.distributions and page is not html:
                        fund.distributions = parse_distributions(html)
                    for _d in fund.distributions:
                        _d.setdefault("source", "issuer")
                    if not fund.distributions:
                        fund.distributions = fetch_dividendhistory(fund)
                    if fund.distributions:
                        fund.stats["distribution_source"] = fund.distributions[0].get("source", "issuer")
                        fund.stats.update(distribution_summary(fund.distributions))
                        log.info("  -> %d distributions, latest %s",
                                 len(fund.distributions), fund.distributions[0]["ex_date"])
                except Exception as exc:  # noqa: BLE001
                    log.warning("  -> distributions FAILED: %s", exc)
            if not fund.holdings:
                fund.error = "parser ran but returned no holdings"
                # keep yesterday's holdings rather than losing them, but the
                # fresh price and distributions above still stand
                prev = previous.get(fund.ticker) or {}
                if prev.get("holdings"):
                    fund.holdings = prev["holdings"]
                    fund.stale = True
        except Exception as exc:  # noqa: BLE001 — log and continue, don't kill the whole run
            fund.fetched_ok = False
            fund.error = str(exc)
            log.warning("  -> FAILED: %s", exc)
            log.warning("  -> %s", describe_response(html))
            carry_forward(fund, previous)

        # A blocked issuer throws before the distribution step above ever runs,
        # so try the fallback out here too — Evolve's holdings 403 shouldn't also
        # cost us their payment history, which is public elsewhere.
        if not fund.holdings:
            _seed = SEED.get(fund.ticker)
            if _seed and _seed.get("holdings"):
                fund.holdings = _seed["holdings"]
                fund.stats["holdings_source"] = _seed.get("source", "manual capture")
                fund.stats["holdings_captured"] = _seed.get("captured", "")
                log.info("  -> %d holdings from seed file", len(fund.holdings))

        if not fund.distributions:
            fund.distributions = fetch_dividendhistory(fund)
            if fund.distributions:
                fund.stats["distribution_source"] = fund.distributions[0].get("source", "issuer")
                fund.stats.update(distribution_summary(fund.distributions))
        time.sleep(DELAY_BETWEEN_REQUESTS)
    return registry


def attach_price_and_yield(registry: list) -> None:
    """Give every fund the same yield measure.

    Issuers publish yields on incompatible bases — a 30-day SEC yield, an
    annualised distribution yield and a trailing twelve-month yield are three
    different numbers, and only 86 of 120 funds publish any at all. We keep the
    issuer's figure (labelled, on the fund page) but also compute one
    comparable number for everyone: twelve months of actual distributions over
    the latest close. Prices come from data/prices, which covers all 120.
    """
    price_dir = Path("data/prices")
    if not price_dir.exists():
        log.warning("no price files — skipping yield computation")
        return
    done = 0
    for fund in registry:
        csv_path = price_dir / f"{fund.ticker}.csv"
        if not csv_path.exists():
            continue
        try:
            last = csv_path.read_text().strip().splitlines()[-1]
            day, close = last.split(",")[0], float(last.split(",")[1])
        except Exception:  # noqa: BLE001
            continue
        if close <= 0:
            continue
        fund.stats["price"] = round(close, 4)
        fund.stats["price_date"] = day
        ttm = fund.stats.get("ttm_total")
        if ttm:
            fund.stats["yield_ttm"] = round(ttm / close * 100, 2)
            done += 1

        # Total return over the same window, from the adjusted series (which
        # already reinvests distributions). A very high yield paired with a
        # weak total return means the distribution is largely coming out of
        # capital — the single most misleading thing about these funds, so the
        # site needs both numbers side by side, not the yield alone.
        try:
            rows = [ln.split(",") for ln in csv_path.read_text().strip().splitlines()[1:]]
            series = [(r[0], float(r[1])) for r in rows if len(r) >= 2]
            if len(series) > 30:
                cutoff = (datetime.now(timezone.utc).date() - timedelta(days=365)).isoformat()
                past = [p for d0, p in series if d0 <= cutoff]
                start = past[-1] if past else series[0][1]
                if start > 0:
                    fund.stats["total_return_1y"] = round((close / start - 1) * 100, 2)
                    fund.stats["return_window"] = "1y" if past else "since inception"
        except Exception:  # noqa: BLE001
            pass
    log.info("Computed a trailing-12-month yield for %d funds", done)


class PushRejected(RuntimeError):
    """The scrape ran but the data never landed.

    This used to be logged and swallowed. The importer spent days
    rejecting every push while the workflow reported success, so the site
    served stale data and nothing said so. A push that does not land is a
    failed run.
    """


def push_to_supabase(payload: dict) -> None:
    """Send the scraped funds to Supabase.

    The site reads fund data through an Edge Function that checks the caller's
    token, so what a visitor can see is decided on the server. That only holds
    if Supabase has the data, hence this push. The local JSON stays as a
    working artefact for the run; it is no longer what the site serves.

    Authenticated with a shared phrase, not a database key: a leaked push key
    can only add fund data, while a service role key could read or delete
    anything. A missing secret warns loudly rather than crashing the run.
    """
    key = os.environ.get("LICENTIA_PUSH_KEY")
    if not key:
        log.warning("LICENTIA_PUSH_KEY not set - scraped data was not pushed to Supabase")
        return

    url = "https://sopzbiuwakowbuqgwpmg.supabase.co/functions/v1/import-funds"
    try:
        resp = requests.post(url, json={"key": key, "funds": payload}, timeout=180)
    except Exception as exc:  # noqa: BLE001
        raise PushRejected(f"could not reach the importer: {exc}") from exc

    if resp.status_code != 200:
        raise PushRejected(
            f"importer returned HTTP {resp.status_code}: {resp.text[:300]}")

    imported = resp.json().get("imported")
    log.info("Pushed %s funds to Supabase", imported)
    if not imported:
        raise PushRejected("importer accepted the request but wrote no rows")


def write_output(registry: list[Fund], path: Path = OUTPUT_PATH,
                 merge: bool = False) -> None:
    """When only part of the registry was scraped (--only), merge into the
    existing file instead of replacing it. Writing a filtered run straight out
    would silently delete every fund we didn't ask for."""
    path.parent.mkdir(parents=True, exist_ok=True)
    payload = {}
    if merge and path.exists():
        try:
            payload = json.loads(path.read_text())
            log.info("Merging %d scraped funds into the existing %d in %s",
                     len(registry), len(payload), path)
        except Exception as exc:  # noqa: BLE001
            log.warning("Could not read existing %s (%s) — writing fresh", path, exc)
            payload = {}

    for fund in registry:
        payload[fund.ticker] = {
            "name": fund.name,
            "issuer": fund.issuer,
            "region": fund.region,
            "holdings": fund.holdings,
            "stats": fund.stats,
            "distributions": fund.distributions,
            "as_of": fund.as_of,
            "stale": fund.stale,
            "fetched_ok": fund.fetched_ok,
            "error": fund.error,
        }
    path.write_text(json.dumps(payload, indent=2))
    # The price job needs to know which tickers exist and which market each
    # trades in. That is not worth gating, and the fund data itself no longer
    # lives in the repo, so leave just the list behind.
    tickers = {t: {"region": v.get("region"), "name": v.get("name")}
                for t, v in payload.items()}
    Path("data/tickers.json").write_text(json.dumps(tickers, indent=2, sort_keys=True))
    log.info("Wrote data/tickers.json — %d tickers", len(tickers))
    push_to_supabase(payload)
    # Coverage per issuer, so a parser that quietly stops working shows up in
    # the next run rather than weeks later when someone notices a total looks
    # wrong. Every failure this project has had would have been visible here.
    by_issuer: dict = {}
    for f in registry:
        row = by_issuer.setdefault(f.issuer, {"n": 0, "h": 0, "a": 0, "d": 0})
        row["n"] += 1
        if f.holdings:
            row["h"] += 1
        if f.stats.get("aum_musd"):
            row["a"] += 1
        if f.distributions:
            row["d"] += 1
    log.info("Coverage by issuer (holdings / size / distributions):")
    for name in sorted(by_issuer):
        r = by_issuer[name]
        flag = "" if r["h"] == r["n"] else "   <-- holdings gap"
        log.info("  %-32s %3d funds   %3d/%d holdings   %3d/%d size   %3d/%d dist%s",
                 name, r["n"], r["h"], r["n"], r["a"], r["n"], r["d"], r["n"], flag)

    ok = sum(1 for f in registry if f.fetched_ok)
    with_aum = sum(1 for f in registry if f.stats.get("aum_musd"))
    with_dist = sum(1 for f in registry if f.distributions)
    carried = [f.ticker for f in registry if f.stale and f.holdings]
    empty = [f.ticker for f in registry if not f.holdings]
    if carried:
        log.warning("Carried forward (issuer unreachable today): %s", ", ".join(carried))
    if empty:
        log.warning("No data at all: %s", ", ".join(empty))
    log.info("Wrote %s — %d/%d fetched live, %d carried forward, %d with AUM, %d with distributions",
             path, ok, len(registry), len(carried), with_aum, with_dist)


def select(registry: list[Fund], only: str) -> list[Fund]:
    """--only accepts tickers, issuer names or parser keys, comma separated and
    case-insensitive: --only=evolve,hamilton / --only=HMAX,BANK. Verifying a
    one-issuer fix shouldn't mean re-scraping 120 funds across every issuer."""
    wanted = {w.strip().lower() for w in only.split(",") if w.strip()}
    picked = [f for f in registry
              if f.ticker.lower() in wanted
              or f.parser.lower() in wanted
              or f.issuer.lower() in wanted
              or any(w in f.issuer.lower() for w in wanted)]
    missing = wanted - {f.ticker.lower() for f in picked} \
                     - {f.parser.lower() for f in picked} \
                     - {f.issuer.lower() for f in picked}
    unmatched = [w for w in missing
                 if not any(w in f.issuer.lower() for f in picked)]
    if unmatched:
        log.warning("--only had no match for: %s", ", ".join(sorted(unmatched)))
    return picked


if __name__ == "__main__":
    import argparse

    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--only", default="",
                    help="comma-separated tickers, issuers or parser keys to scrape")
    args = ap.parse_args()

    registry = FUND_REGISTRY
    filtered = bool(args.only)
    if filtered:
        registry = select(registry, args.only)
        if not registry:
            raise SystemExit(f"--only={args.only!r} matched no funds")
        log.info("Scraping %d of %d funds: %s", len(registry), len(FUND_REGISTRY),
                 ", ".join(f.ticker for f in registry))

    results = run(registry)
    attach_price_and_yield(results)
    write_output(results, merge=filtered)



# ---------------------------------------------------------------------------
# Example: .github/workflows/update-data.yml
# Save this as a separate file in your repo to automate the weekly/daily run.
# ---------------------------------------------------------------------------
GITHUB_ACTIONS_EXAMPLE = """
name: Update ETF holdings data
on:
  schedule:
    - cron: '0 13 * * 1'   # every Monday 13:00 UTC — adjust to daily if you want
  workflow_dispatch: {}     # lets you trigger it manually from the Actions tab

jobs:
  scrape:
    runs-on: ubuntu-latest
    steps:
      - uses: actions/checkout@v4
      - uses: actions/setup-python@v5
        with:
          python-version: '3.12'
      - run: pip install requests beautifulsoup4 lxml
      - run: python scraper.py
      - name: Commit updated data
        run: |
          git config user.name "ledger-bot"
          git config user.email "bot@users.noreply.github.com"
          git add data/funds.json
          git diff --quiet --cached || git commit -m "Update fund holdings data"
          git push
"""
